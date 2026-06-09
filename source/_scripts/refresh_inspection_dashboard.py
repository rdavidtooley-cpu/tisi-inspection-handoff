#!/usr/bin/env python3
"""
Inspection Intel - NDT Dashboard Refresh Script
Fetches equity data via yfinance, O&G/customer industry data from FRED,
generates Excel workbook, and injects data into HTML dashboards.

Usage: python3 refresh_inspection_dashboard.py [--dry-run]
Dependencies: pip install yfinance openpyxl
"""

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json
import os
import sys
import csv
import re
import logging
import statistics
import traceback
from datetime import datetime, timedelta
from pathlib import Path
import time
import ssl
import certifi
from urllib.request import urlopen, Request
from urllib.error import URLError

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_DIR = Path(os.environ.get('PROJECT_DIR', str(SCRIPT_DIR.parent)))
DASHBOARD_DIR = Path(os.environ.get('DASHBOARD_DIR', str(PROJECT_DIR / 'Dashboard')))
FINANCIALS_DIR = PROJECT_DIR / 'Financials'
INDUSTRY_DIR = PROJECT_DIR / 'Industry_Data'
LOG_DIR = SCRIPT_DIR / 'logs'

MARKET_DATA_FILE = DASHBOARD_DIR / 'market_data.json'
PRICE_HISTORY_FILE = DASHBOARD_DIR / 'price_history.json'
INDUSTRY_INDICATORS_FILE = DASHBOARD_DIR / 'industry_indicators.json'
EXCEL_FILE = FINANCIALS_DIR / 'TIC_NDT_Market_Data.xlsx'

HISTORY_DAYS = 365
BIG_MOVER_THRESHOLD = 3.0  # Lower threshold for smaller-cap universe

USER_AGENT = 'InspectionIntel RobertTooley __ADMIN_EMAIL__'

# ── Ticker Universe ──
# 21 public tickers across 5 categories
# Includes international companies on SIX, Euronext, LSE exchanges

TICKER_UNIVERSE = {
    # NDT Services (US-listed)
    'MG':       {'name': 'Mistras Group',      'category': 'NDT Services',  'hq': 'USA',         'exchange': 'NYSE'},
    'TISI':     {'name': 'Team Inc',           'category': 'NDT Services',  'hq': 'USA',         'exchange': 'NYSE'},
    'TIC':      {'name': 'Acuren Group',       'category': 'NDT Services',  'hq': 'Canada',      'exchange': 'NYSE'},
    'OII':      {'name': 'Oceaneering Intl',   'category': 'NDT Services',  'hq': 'USA',         'exchange': 'NYSE'},
    'XPRO':     {'name': 'Expro Group',        'category': 'NDT Services',  'hq': 'UK',          'exchange': 'NYSE'},
    # Global NDT (international exchanges)
    'BVI.PA':   {'name': 'Bureau Veritas',     'category': 'Global NDT',    'hq': 'France',       'exchange': 'Euronext'},
    'ITRK.L':   {'name': 'Intertek Group',     'category': 'Global NDT',    'hq': 'UK',           'exchange': 'LSE'},
    'COTN.SW':  {'name': 'Comet Group',        'category': 'Global NDT',    'hq': 'Switzerland',  'exchange': 'SIX'},
    'SGSN.SW':  {'name': 'SGS SA',             'category': 'Global NDT',    'hq': 'Switzerland',  'exchange': 'SIX'},
    # NDT Adjacent
    'TRNS':     {'name': 'Transcat',           'category': 'NDT Adjacent',  'hq': 'USA',         'exchange': 'NASDAQ'},
    'THR':      {'name': 'Thermon Group',      'category': 'NDT Adjacent',  'hq': 'USA',         'exchange': 'NYSE'},
    # Flow Control (valves, actuators, flow products)
    'FLS':      {'name': 'Flowserve',          'category': 'Flow Control',  'hq': 'USA',         'exchange': 'NYSE'},
    'ROR.L':    {'name': 'Rotork',             'category': 'Flow Control',  'hq': 'UK',          'exchange': 'LSE'},
    'IMI.L':    {'name': 'IMI plc',            'category': 'Flow Control',  'hq': 'UK',          'exchange': 'LSE'},
    'SPX.L':    {'name': 'Spirax Group',       'category': 'Flow Control',  'hq': 'UK',          'exchange': 'LSE'},
    'WEIR.L':   {'name': 'Weir Group',         'category': 'Flow Control',  'hq': 'UK',          'exchange': 'LSE'},
    'WHD':      {'name': 'Cactus Inc',         'category': 'Flow Control',  'hq': 'USA',         'exchange': 'NYSE'},
    # Mechanical & On-Site Services (pipeline, infra, industrial services)
    'MTRX':     {'name': 'Matrix Service',     'category': 'Mech. & On-Site Services', 'hq': 'USA', 'exchange': 'NASDAQ'},
    'PRIM':     {'name': 'Primoris Services',  'category': 'Mech. & On-Site Services', 'hq': 'USA', 'exchange': 'NYSE'},
    'MTZ':      {'name': 'MasTec',             'category': 'Mech. & On-Site Services', 'hq': 'USA', 'exchange': 'NYSE'},
    'CLH':      {'name': 'Clean Harbors',      'category': 'Mech. & On-Site Services', 'hq': 'USA', 'exchange': 'NYSE'},
    'FET':      {'name': 'Forum Energy Tech',  'category': 'Mech. & On-Site Services', 'hq': 'USA', 'exchange': 'NYSE'},
}

CATEGORY_ORDER = ['NDT Services', 'Global NDT', 'NDT Adjacent', 'Flow Control', 'Mech. & On-Site Services']
CATEGORY_COLORS = {
    'NDT Services':             '#4fc3f7',  # blue
    'Global NDT':               '#ffd54f',  # gold
    'NDT Adjacent':             '#ba68c8',  # purple
    'Flow Control':             '#81c784',  # green
    'Mech. & On-Site Services': '#ff8a65',  # orange
}

# Two thematic baskets — must be kept in sync with TICKER_UNIVERSE composition.
# Inspection-11 = TIC/NDT pure-play (NDT Services + Global NDT + NDT Adjacent).
# Flow & MOS-11 = adjacent industrial services peers (Flow Control + Mech. & On-Site Services).
INSPECTION_11_TICKERS = ['MG', 'TISI', 'TIC', 'OII', 'XPRO', 'BVI.PA', 'ITRK.L', 'COTN.SW', 'SGSN.SW', 'TRNS', 'THR']
FLOW_MOS_TICKERS = ['FLS', 'ROR.L', 'IMI.L', 'SPX.L', 'WEIR.L', 'WHD', 'MTRX', 'PRIM', 'MTZ', 'CLH', 'FET']

# Display ticker mapping (for dashboard display vs yfinance fetch)
DISPLAY_TICKER = {
    'BVI.PA':  'BVI',
    'ITRK.L':  'ITRK',
    'COTN.SW': 'COTN',
    'SGSN.SW': 'SGSN',
    'ROR.L':   'ROR',
    'IMI.L':   'IMI',
    'SPX.L':   'SPX',
    'WEIR.L':  'WEIR',
}

# FRED series for customer industry data (free, no API key)
FRED_SERIES = {
    # Energy prices
    'DCOILWTICO':   'WTI Crude Oil ($/bbl)',
    'DCOILBRENTEU': 'Brent Crude Oil ($/bbl)',
    'DHHNGSP':      'Henry Hub Natural Gas ($/MMBtu)',
    'DHOILNYH':     'No. 2 Heating Oil — NY Harbor ($/gal)',
    'GASREGW':      'US Regular Gasoline ($/gal)',
    # Manufacturing & Industrial
    'INDPRO':       'Industrial Production Index',
    'MANEMP':       'Manufacturing Employment (thousands)',
    'IPMINE':       'Mining Industrial Production Index',
    'DGORDER':      'Durable Goods Orders ($M)',
    # Construction & Infrastructure
    'TLRESCONS':    'Total Construction Spending ($M)',
    'TLNRESCONS':   'Nonresidential Construction ($M)',
    'TLPWRCONS':    'Power Construction Spending ($M)',
    'PBHWYCONS':    'Highway & Street Construction ($M)',
    'TLMFGCONS':    'Manufacturing Construction ($M)',
    # Oil & Gas Production / Drilling
    'IPN213111N':   'Drilling Oil & Gas Wells Index',
    'CAPUTLG211S':  'O&G Extraction Capacity Util. (%)',
    'CAPUTLG324S':  'Refinery Capacity Utilization (%)',
    'IPG211S':      'Oil & Gas Extraction Production Index',
    # Aerospace & Defense
    'ADEFNO':       'Defense Capital Goods Orders ($M)',
    'UNAPNO':       'Commercial Aircraft Orders ($M)',
    # Nuclear & Power
    'IPN221113S':   'Nuclear Power Generation Index',
    'CAPUTLG2211S': 'Electric Utility Capacity Util. (%)',
    # Additional NDT demand proxies
    'A32SNO':       'Fabricated Metal Products Orders ($M)',
    'IPG325S':      'Chemical Manufacturing Index',
    'NEWORDER':     'Core Capital Goods Orders ($M)',
    # Macro indicators
    'MCUMFN':       'Manufacturing Capacity Utilization (%)',
    'PERMIT':       'Building Permits (thousands)',
    'UNRATE':       'Unemployment Rate (%)',
}

CUSTOMER_INDUSTRY_WEIGHTS = {
    'Oil & Gas': 30.0,
    'Power Generation': 15.0,
    'Aerospace & Defense': 15.0,
    'Manufacturing': 12.0,
    'Marine & Offshore': 10.0,
    'Infrastructure': 10.0,
    'Other': 8.0,
}

# =============================================================================
# LOGGING
# =============================================================================

def setup_logging():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f'dashboard_refresh_{datetime.now().strftime("%Y-%m-%d")}.log'
    logger = logging.getLogger('inspection_dashboard')
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        fh = logging.FileHandler(log_file)
        fh.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        fmt = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        fh.setFormatter(fmt)
        ch.setFormatter(fmt)
        logger.addHandler(fh)
        logger.addHandler(ch)
    return logger


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def safe_div(a, b):
    if a is not None and b is not None and b != 0:
        return a / b
    return None


def fmt_val(value, fallback='N/A'):
    return value if value is not None else fallback


def linear_trend(prices):
    """Annualized linear regression slope as % of mean price."""
    n = len(prices)
    if n < 10:
        return None
    x = list(range(n))
    mean_x = sum(x) / n
    mean_y = sum(prices) / n
    if mean_y == 0:
        return None
    numerator = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, prices))
    denominator = sum((xi - mean_x) ** 2 for xi in x)
    if denominator == 0:
        return None
    slope_per_day = numerator / denominator
    return round(slope_per_day * 252 / mean_y * 100, 1)


def classify_mcap_tier(market_cap_b):
    if market_cap_b is None or market_cap_b <= 0:
        return 'Unknown'
    if market_cap_b >= 200:
        return 'Mega-Cap (>$200B)'
    elif market_cap_b >= 10:
        return 'Large-Cap ($10B-$200B)'
    elif market_cap_b >= 2:
        return 'Mid-Cap ($2B-$10B)'
    elif market_cap_b >= 0.3:
        return 'Small-Cap ($300M-$2B)'
    elif market_cap_b >= 0.05:
        return 'Micro-Cap ($50M-$300M)'
    return 'Nano-Cap (<$50M)'


# =============================================================================
# YFINANCE DATA FETCHING
# =============================================================================

def fetch_fx_rates(logger):
    """Fetch exchange rates for AUD, GBP, CHF, EUR → USD."""
    rates = {'AUD': 0.65, 'GBP': 1.27, 'CHF': 1.10, 'EUR': 1.08}  # fallback defaults
    for ccy, pair in [('AUD', 'AUDUSD=X'), ('GBP', 'GBPUSD=X'), ('CHF', 'CHFUSD=X'), ('EUR', 'EURUSD=X')]:
        try:
            fx = yf.Ticker(pair)
            rate = fx.info.get('regularMarketPrice') or fx.info.get('previousClose')
            if rate:
                rates[ccy] = float(rate)
                logger.info(f'  {ccy}/USD rate: {rates[ccy]:.4f}')
        except Exception as e:
            logger.warning(f'  Failed to fetch {ccy}/USD rate: {e}, using fallback {rates[ccy]}')
    return rates


def fetch_ticker_data(ticker, meta, fx_rates, logger):
    """Fetch live data for a single ticker via yfinance."""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        # Price with cascade fallback
        price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
        if price is None:
            price = 0

        market_cap_raw = info.get('marketCap', 0) or 0
        ev_raw = info.get('enterpriseValue', 0) or 0

        # Currency handling
        currency = info.get('currency', 'USD')
        price_usd = price
        mcap_usd = market_cap_raw
        if currency == 'AUD' and fx_rates.get('AUD'):
            price_usd = price * fx_rates['AUD']
            mcap_usd = market_cap_raw * fx_rates['AUD']
        elif currency == 'GBp' and fx_rates.get('GBP'):
            # GBp = pence sterling — divide by 100 to get GBP, then convert to USD
            price_usd = (price / 100) * fx_rates['GBP']
            mcap_usd = market_cap_raw * fx_rates['GBP']  # marketCap already in GBP from yfinance
        elif currency == 'GBP' and fx_rates.get('GBP'):
            price_usd = price * fx_rates['GBP']
            mcap_usd = market_cap_raw * fx_rates['GBP']
        elif currency == 'CHF' and fx_rates.get('CHF'):
            price_usd = price * fx_rates['CHF']
            mcap_usd = market_cap_raw * fx_rates['CHF']
        elif currency == 'EUR' and fx_rates.get('EUR'):
            price_usd = price * fx_rates['EUR']
            mcap_usd = market_cap_raw * fx_rates['EUR']

        data = {
            'ticker': ticker,
            'display_ticker': DISPLAY_TICKER.get(ticker, ticker),
            'company': info.get('longName') or meta['name'],
            'name': meta['name'],
            'category': meta['category'],
            'hq': meta['hq'],
            'exchange': meta['exchange'],
            'price': round(price, 2),
            'price_usd': round(price_usd, 2),
            'currency': currency,
            'market_cap_b': round(mcap_usd / 1e9, 3) if mcap_usd else 0,
            'enterprise_value_b': round(ev_raw / 1e9, 3) if ev_raw else None,
            'mcap_tier': classify_mcap_tier(mcap_usd / 1e9 if mcap_usd else 0),

            # Valuation
            'pe_trailing': info.get('trailingPE'),
            'pe_forward': info.get('forwardPE'),
            'ev_ebitda': info.get('enterpriseToEbitda'),
            'ev_revenue': info.get('enterpriseToRevenue'),
            'price_to_book': info.get('priceToBook'),
            'price_to_sales': info.get('priceToSalesTrailing12Months'),
            'eps_trailing': info.get('trailingEps'),
            'eps_forward': info.get('forwardEps'),
            'book_value': info.get('bookValue'),
            'revenue_per_share': info.get('revenuePerShare'),

            # Growth
            'revenue_growth': info.get('revenueGrowth'),
            'earnings_growth': info.get('earningsGrowth'),
            'quarterly_earnings_growth': info.get('earningsQuarterlyGrowth'),

            # Profitability
            'gross_margins': info.get('grossMargins'),
            'operating_margins': info.get('operatingMargins'),
            'profit_margins': info.get('profitMargins'),
            'ebitda_margins': info.get('ebitdaMargins'),
            'roe': info.get('returnOnEquity'),
            'roa': info.get('returnOnAssets'),
            'revenue_b': round(info.get('totalRevenue', 0) / 1e9, 3) if info.get('totalRevenue') else None,
            'ebitda_b': round(info.get('ebitda', 0) / 1e9, 3) if info.get('ebitda') else None,
            'net_income_b': round(info.get('netIncomeToCommon', 0) / 1e9, 3) if info.get('netIncomeToCommon') else None,
            'operating_cashflow_b': round(info.get('operatingCashflow', 0) / 1e9, 3) if info.get('operatingCashflow') else None,
            'free_cashflow_b': round(info.get('freeCashflow', 0) / 1e9, 3) if info.get('freeCashflow') else None,

            # Financial Health
            'total_debt_b': round(info.get('totalDebt', 0) / 1e9, 3) if info.get('totalDebt') else None,
            'total_cash_b': round(info.get('totalCash', 0) / 1e9, 3) if info.get('totalCash') else None,
            'debt_to_equity': info.get('debtToEquity'),
            'current_ratio': info.get('currentRatio'),
            'quick_ratio': info.get('quickRatio'),

            # Risk
            'beta': info.get('beta'),

            # Analyst
            'target_high': info.get('targetHighPrice'),
            'target_low': info.get('targetLowPrice'),
            'target_mean': info.get('targetMeanPrice'),
            'target_median': info.get('targetMedianPrice'),
            'recommendation_key': info.get('recommendationKey'),
            'recommendation_mean': info.get('recommendationMean'),
            'num_analysts': info.get('numberOfAnalystOpinions'),

            # Short interest
            'shares_short': info.get('sharesShort'),
            'short_ratio': info.get('shortRatio'),
            'short_pct_float': info.get('shortPercentOfFloat'),

            # Ownership
            'held_pct_insiders': info.get('heldPercentInsiders'),
            'held_pct_institutions': info.get('heldPercentInstitutions'),

            # Dividends
            'dividend_yield': info.get('dividendYield'),
            'dividend_rate': info.get('dividendRate'),
            'payout_ratio': info.get('payoutRatio'),

            # Technical
            'fifty_two_week_high': info.get('fiftyTwoWeekHigh', 0),
            'fifty_two_week_low': info.get('fiftyTwoWeekLow', 0),
            'fifty_day_avg': info.get('fiftyDayAverage'),
            'two_hundred_day_avg': info.get('twoHundredDayAverage'),
            'avg_volume_10d': info.get('averageVolume10days'),
            'avg_volume_3mo': info.get('averageVolume'),
            'day_high': info.get('dayHigh'),
            'day_low': info.get('dayLow'),
            'open_price': info.get('open'),
            'previous_close': info.get('previousClose'),
            'daily_change': info.get('regularMarketChange'),
            'daily_change_pct': info.get('regularMarketChangePercent'),
            'change_pct': info.get('regularMarketChangePercent'),

            # Company info
            'employees': info.get('fullTimeEmployees'),
            'industry_yf': info.get('industry'),
            'sector_yf': info.get('sector'),
            'country': info.get('country'),
            'website': info.get('website'),

            # Timestamps
            'most_recent_quarter': None,
            'earnings_timestamp': None,
            'next_earnings_date': None,
        }

        # Convert Unix timestamps
        for field, key in [('most_recent_quarter', 'mostRecentQuarter'),
                           ('earnings_timestamp', 'earningsTimestamp')]:
            raw_ts = info.get(key)
            if raw_ts and isinstance(raw_ts, (int, float)):
                try:
                    data[field] = datetime.fromtimestamp(raw_ts).strftime('%Y-%m-%d')
                except (OSError, ValueError):
                    pass

        # Get next earnings date from calendar (more reliable than earningsTimestamp)
        try:
            cal = stock.calendar
            if cal and 'Earnings Date' in cal:
                ed = cal['Earnings Date']
                if isinstance(ed, list) and len(ed) > 0:
                    data['next_earnings_date'] = str(ed[0])
                elif ed:
                    data['next_earnings_date'] = str(ed)
        except Exception:
            pass

        # Earnings history (actual vs estimate, last 12 quarters)
        try:
            ed_df = stock.earnings_dates
            if ed_df is not None and not ed_df.empty:
                earnings_hist = []
                for dt_idx, erow in ed_df.iterrows():
                    reported = erow.get('Reported EPS')
                    estimated = erow.get('EPS Estimate')
                    surprise = erow.get('Surprise(%)')
                    if reported is not None and not (isinstance(reported, float) and (reported != reported)):
                        earnings_hist.append({
                            'date': dt_idx.strftime('%Y-%m-%d') if hasattr(dt_idx, 'strftime') else str(dt_idx)[:10],
                            'eps_actual': round(float(reported), 2) if reported == reported else None,
                            'eps_estimate': round(float(estimated), 2) if estimated is not None and estimated == estimated else None,
                            'surprise_pct': round(float(surprise), 2) if surprise is not None and surprise == surprise else None,
                        })
                data['earnings_history'] = earnings_hist[:12]
        except Exception:
            data['earnings_history'] = []

        # Convert native-currency fields to USD (analyst targets, technicals)
        # yfinance returns these in the stock's native currency
        if currency != 'USD' and price > 0 and price_usd > 0:
            fx_mult_per_unit = price_usd / price  # USD per 1 unit of native currency
            native_fields = [
                'target_high', 'target_low', 'target_mean', 'target_median',
                'fifty_two_week_high', 'fifty_two_week_low',
                'fifty_day_avg', 'two_hundred_day_avg',
                'day_high', 'day_low', 'open_price', 'previous_close',
                'eps_trailing', 'eps_forward', 'book_value', 'revenue_per_share',
            ]
            for field in native_fields:
                if data.get(field) is not None:
                    data[field] = round(data[field] * fx_mult_per_unit, 2)

        # % off 52-week high (now both in USD)
        if data['price_usd'] and data['fifty_two_week_high']:
            data['pct_off_high'] = round((data['price_usd'] - data['fifty_two_week_high']) / data['fifty_two_week_high'] * 100, 1)
        else:
            data['pct_off_high'] = None

        # % vs moving averages (now both in USD)
        if data['price_usd'] and data['fifty_day_avg']:
            data['vs_50dma_pct'] = round((data['price_usd'] - data['fifty_day_avg']) / data['fifty_day_avg'] * 100, 1)
        else:
            data['vs_50dma_pct'] = None
        if data['price_usd'] and data['two_hundred_day_avg']:
            data['vs_200dma_pct'] = round((data['price_usd'] - data['two_hundred_day_avg']) / data['two_hundred_day_avg'] * 100, 1)
        else:
            data['vs_200dma_pct'] = None

        # Analyst upside (now both in USD)
        if data['target_mean'] and data['price_usd'] and data['price_usd'] > 0:
            data['upside_pct'] = round((data['target_mean'] - data['price_usd']) / data['price_usd'] * 100, 1)
        else:
            data['upside_pct'] = None

        # 1Y and YTD changes + trend + daily price series (use local currency price for both sides)
        try:
            hist = stock.history(period='1y')
            if len(hist) > 0:
                price_1y_ago = float(hist.iloc[0]['Close'])
                if price_1y_ago > 0:
                    data['change_1y'] = round((price - price_1y_ago) / price_1y_ago * 100, 1)
                else:
                    data['change_1y'] = 0
                data['trend_1y'] = linear_trend(hist['Close'].tolist())

                # Build daily price series for sparkline charts (in USD)
                # Determine FX multiplier for converting local prices to USD
                fx_mult = 1.0
                if currency == 'GBp' and fx_rates.get('GBP'):
                    fx_mult = fx_rates['GBP'] / 100
                elif currency == 'GBP' and fx_rates.get('GBP'):
                    fx_mult = fx_rates['GBP']
                elif currency == 'AUD' and fx_rates.get('AUD'):
                    fx_mult = fx_rates['AUD']
                elif currency == 'CHF' and fx_rates.get('CHF'):
                    fx_mult = fx_rates['CHF']
                elif currency == 'EUR' and fx_rates.get('EUR'):
                    fx_mult = fx_rates['EUR']

                daily_prices = []
                for dt, row in hist.iterrows():
                    close_val = float(row['Close'])
                    daily_prices.append({
                        'd': dt.strftime('%Y-%m-%d'),
                        'p': round(close_val * fx_mult, 2)
                    })
                data['price_series'] = daily_prices
            else:
                data['change_1y'] = 0
                data['trend_1y'] = None
                data['price_series'] = []
        except Exception:
            data['change_1y'] = 0
            data['trend_1y'] = None
            data['price_series'] = []

        try:
            today = datetime.now()
            ytd_start = datetime(today.year, 1, 1)
            hist_ytd = stock.history(start=ytd_start)
            if len(hist_ytd) > 0:
                price_ytd = float(hist_ytd.iloc[0]['Close'])
                if price_ytd > 0:
                    data['change_ytd'] = round((price - price_ytd) / price_ytd * 100, 1)
                else:
                    data['change_ytd'] = 0
            else:
                data['change_ytd'] = 0
        except Exception:
            data['change_ytd'] = 0

        # Institutional holders (top 10)
        try:
            inst_holders = stock.institutional_holders
            shares_out = info.get('sharesOutstanding')
            if inst_holders is not None and not inst_holders.empty:
                holders_list = []
                for _, row in inst_holders.head(10).iterrows():
                    shares_val = int(row['Shares']) if row.get('Shares') is not None and row['Shares'] == row['Shares'] else None
                    # Try % Out column first, fall back to computing from shares/outstanding
                    pct_val = None
                    if row.get('% Out') is not None and row['% Out'] == row['% Out']:
                        pct_val = round(float(row['% Out']) * 100, 2)
                    elif shares_val and shares_out and shares_out > 0:
                        pct_val = round(shares_val / shares_out * 100, 2)
                    holders_list.append({
                        'holder': str(row.get('Holder', '')),
                        'shares': shares_val,
                        'date_reported': str(row['Date Reported'])[:10] if row.get('Date Reported') is not None else None,
                        'pct_out': pct_val,
                        'value': round(float(row['Value']) / 1e6, 1) if row.get('Value') is not None and row['Value'] == row['Value'] else None,
                    })
                data['institutional_holders'] = holders_list
            else:
                data['institutional_holders'] = []
        except Exception:
            data['institutional_holders'] = []

        data['timestamp'] = datetime.now().isoformat()
        return data

    except Exception as e:
        logger.error(f'  FAILED {ticker}: {e}')
        return None


def fetch_all_market_data(logger):
    """Fetch data for all tickers in the universe."""
    logger.info(f'Fetching market data for {len(TICKER_UNIVERSE)} tickers...')
    market_data = {}

    # Get exchange rates for foreign tickers
    fx_rates = fetch_fx_rates(logger)

    tickers = list(TICKER_UNIVERSE.keys())
    for i in range(0, len(tickers), 8):
        batch = tickers[i:i+8]
        logger.info(f'  Batch {i//8 + 1}/{(len(tickers)+7)//8}: {" ".join(batch)}')

        # Warm cache
        try:
            yf.download(' '.join(batch), period='1d', group_by='ticker', progress=False, threads=False)
        except Exception:
            pass

        for ticker in batch:
            time.sleep(0.3)  # Prevent DNS thread exhaustion on macOS
            meta = TICKER_UNIVERSE[ticker]
            data = fetch_ticker_data(ticker, meta, fx_rates, logger)
            if data:
                market_data[ticker] = data
                logger.info(f'    {ticker}: ${data["price_usd"]:.2f} MCap=${data["market_cap_b"]:.1f}B ({data["change_1y"]:+.1f}% 1Y)')
            else:
                logger.warning(f'    {ticker}: FAILED')

    logger.info(f'Fetched {len(market_data)}/{len(TICKER_UNIVERSE)} tickers successfully')
    return market_data


def _build_income_records(fin, period_fmt='year'):
    """Build income statement records from a yfinance financials DataFrame.
    period_fmt: 'year' for annual (col.strftime('%Y')), 'quarter' for quarterly ('2024-Q4').
    """
    LINE_ITEMS = ['Total Revenue', 'Cost Of Revenue', 'Gross Profit', 'Operating Income', 'EBITDA', 'Net Income']
    periods = {}
    for col in fin.columns:
        if period_fmt == 'year':
            key = col.strftime('%Y')
        else:
            q = (col.month - 1) // 3 + 1
            key = f"{col.strftime('%Y')}-Q{q}"
        p_data = {}
        for item in LINE_ITEMS:
            if item in fin.index:
                val = fin.loc[item, col]
                p_data[item] = round(float(val) / 1e9, 3) if val == val else None
            else:
                p_data[item] = None
        if p_data.get('Total Revenue') is not None:
            periods[key] = p_data
    if not periods:
        return []
    sorted_keys = sorted(periods.keys())
    records = []
    for i, k in enumerate(sorted_keys):
        rec = {'period': k, **periods[k]}
        rev = periods[k].get('Total Revenue')
        if rev and rev > 0:
            gp = periods[k].get('Gross Profit')
            oi = periods[k].get('Operating Income')
            ni = periods[k].get('Net Income')
            eb = periods[k].get('EBITDA')
            rec['gross_margin'] = round(gp / rev * 100, 1) if gp is not None else None
            rec['operating_margin'] = round(oi / rev * 100, 1) if oi is not None else None
            rec['net_margin'] = round(ni / rev * 100, 1) if ni is not None else None
            rec['ebitda_margin'] = round(eb / rev * 100, 1) if eb is not None else None
        if i > 0:
            prev_rev = periods[sorted_keys[i - 1]].get('Total Revenue')
            if prev_rev and prev_rev > 0 and rev:
                rec['revenue_growth_yoy'] = round((rev - prev_rev) / prev_rev * 100, 1)
            else:
                rec['revenue_growth_yoy'] = None
        else:
            rec['revenue_growth_yoy'] = None
        records.append(rec)
    return records


def _build_balance_sheet_records(bs, period_fmt='year'):
    """Build balance sheet records from a yfinance balance sheet DataFrame."""
    BS_ITEMS = [
        'Total Assets', 'Total Liabilities Net Minority Interest',
        'Stockholders Equity', 'Total Debt',
        'Cash And Cash Equivalents', 'Current Assets', 'Current Liabilities',
        'Accounts Receivable', 'Accounts Payable',
    ]
    BS_LABELS = {
        'Total Assets': 'Total Assets',
        'Total Liabilities Net Minority Interest': 'Total Liabilities',
        'Stockholders Equity': 'Stockholders Equity',
        'Total Debt': 'Total Debt',
        'Cash And Cash Equivalents': 'Cash',
        'Current Assets': 'Current Assets',
        'Current Liabilities': 'Current Liabilities',
        'Accounts Receivable': 'Accounts Receivable',
        'Accounts Payable': 'Accounts Payable',
    }
    periods = {}
    for col in bs.columns:
        if period_fmt == 'year':
            key = col.strftime('%Y')
        else:
            q = (col.month - 1) // 3 + 1
            key = f"{col.strftime('%Y')}-Q{q}"
        p_data = {}
        for item in BS_ITEMS:
            label = BS_LABELS[item]
            if item in bs.index:
                val = bs.loc[item, col]
                p_data[label] = round(float(val) / 1e9, 3) if val == val else None
            else:
                p_data[label] = None
        # Compute current ratio
        ca = p_data.get('Current Assets')
        cl = p_data.get('Current Liabilities')
        if ca is not None and cl is not None and cl > 0:
            p_data['Current Ratio'] = round(ca / cl, 2)
        else:
            p_data['Current Ratio'] = None
        # Debt-to-Equity
        td = p_data.get('Total Debt')
        eq = p_data.get('Stockholders Equity')
        if td is not None and eq is not None and eq > 0:
            p_data['Debt to Equity'] = round(td / eq * 100, 1)
        else:
            p_data['Debt to Equity'] = None
        periods[key] = p_data
    if not periods:
        return []
    sorted_keys = sorted(periods.keys())
    return [{'period': k, **periods[k]} for k in sorted_keys]


def _build_cashflow_records(cf, period_fmt='year'):
    """Build cash flow statement records from a yfinance cashflow DataFrame."""
    CF_ITEMS = [
        'Operating Cash Flow', 'Capital Expenditure', 'Free Cash Flow',
        'Depreciation And Amortization',
    ]
    CF_LABELS = {
        'Operating Cash Flow': 'Operating Cash Flow',
        'Capital Expenditure': 'Capital Expenditure',
        'Free Cash Flow': 'Free Cash Flow',
        'Depreciation And Amortization': 'Depreciation & Amortization',
    }
    periods = {}
    for col in cf.columns:
        if period_fmt == 'year':
            key = col.strftime('%Y')
        else:
            q = (col.month - 1) // 3 + 1
            key = f"{col.strftime('%Y')}-Q{q}"
        p_data = {}
        for item in CF_ITEMS:
            label = CF_LABELS[item]
            if item in cf.index:
                val = cf.loc[item, col]
                p_data[label] = round(float(val) / 1e9, 3) if val == val else None
            else:
                p_data[label] = None
        periods[key] = p_data
    if not periods:
        return []
    sorted_keys = sorted(periods.keys())
    return [{'period': k, **periods[k]} for k in sorted_keys]


# =============================================================================
# SEC EDGAR XBRL FALLBACK — fills gaps when yfinance is missing data
# =============================================================================

# CIK registry for SEC EDGAR lookups (US-listed tickers only)
_SEC_CIK_MAP = {}

def _load_sec_cik_map():
    """Load CIK numbers from edgar_company_registry.json."""
    global _SEC_CIK_MAP
    if _SEC_CIK_MAP:
        return _SEC_CIK_MAP
    registry_file = SCRIPT_DIR / 'edgar_company_registry.json'
    if not registry_file.exists():
        return {}
    try:
        with open(registry_file) as f:
            data = json.load(f)
        for co in data.get('companies', []):
            if co.get('cik') and co.get('active', False):
                _SEC_CIK_MAP[co['ticker']] = co['cik']
    except Exception:
        pass
    return _SEC_CIK_MAP


# XBRL tag mappings: each value is a list of tags to try in order
_XBRL_INCOME_TAGS = {
    'Total Revenue': ['RevenueFromContractWithCustomerExcludingAssessedTax', 'Revenues',
                      'SalesRevenueNet', 'SalesRevenueGoodsNet', 'SalesRevenueServicesNet',
                      'RevenueFromContractWithCustomerIncludingAssessedTax'],
    'Cost Of Revenue': ['CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfGoodsSold', 'CostOfServices'],
    'Gross Profit': ['GrossProfit'],
    'Operating Income': ['OperatingIncomeLoss'],
    'Net Income': ['NetIncomeLoss', 'ProfitLoss', 'NetIncomeLossAvailableToCommonStockholdersBasic'],
    'EBITDA': [],  # Not directly in XBRL — computed below
}

_XBRL_BALANCE_TAGS = {
    'Total Assets': ['Assets'],
    'Total Liabilities': ['Liabilities'],
    'Stockholders Equity': ['StockholdersEquity', 'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest'],
    'Total Debt': ['LongTermDebt', 'LongTermDebtAndCapitalLeaseObligations'],
    'Cash': ['CashAndCashEquivalentsAtCarryingValue', 'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents'],
    'Current Assets': ['AssetsCurrent'],
    'Current Liabilities': ['LiabilitiesCurrent'],
    'Accounts Receivable': ['AccountsReceivableNetCurrent', 'ReceivablesNetCurrent'],
    'Accounts Payable': ['AccountsPayableCurrent'],
}

_XBRL_CASHFLOW_TAGS = {
    'Operating Cash Flow': ['NetCashProvidedByUsedInOperatingActivities',
                            'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations'],
    'Capital Expenditure': ['PaymentsToAcquirePropertyPlantAndEquipment'],
    'Depreciation & Amortization': ['DepreciationDepletionAndAmortization', 'DepreciationAndAmortization'],
}

# Extra tags for EBITDA computation
_XBRL_DA_TAGS = ['DepreciationDepletionAndAmortization', 'DepreciationAndAmortization']
_XBRL_INTEREST_TAGS = ['InterestExpense', 'InterestExpenseNonoperating']
_XBRL_TAX_TAGS = ['IncomeTaxExpenseBenefit']


def _fetch_sec_companyfacts(cik, logger):
    """Fetch company facts JSON from SEC EDGAR XBRL API."""
    ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    url = f'https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json'
    req = Request(url, headers={'User-Agent': USER_AGENT, 'Accept-Encoding': 'gzip'})
    try:
        import gzip as _gzip
        with urlopen(req, timeout=30, context=ssl_ctx) as resp:
            raw = resp.read()
            if resp.headers.get('Content-Encoding') == 'gzip':
                raw = _gzip.decompress(raw)
            return json.loads(raw)
    except Exception as e:
        logger.warning(f'    SEC EDGAR fetch failed for CIK {cik}: {e}')
        return None


def _extract_xbrl_values(facts, tag_list, form_filter, frame_pattern, unit='USD'):
    """Extract values from SEC EDGAR facts for a list of candidate XBRL tags.
    Returns dict of {period_key: value_in_billions} for the first tag that has data.
    form_filter: '10-K' or '10-Q'
    frame_pattern: regex to match frame field (e.g. r'^CY\\d{4}$' for annual duration)
    """
    import re as _re
    us_gaap = facts.get('facts', {}).get('us-gaap', {})
    for tag in tag_list:
        tag_data = us_gaap.get(tag)
        if not tag_data:
            continue
        entries = tag_data.get('units', {}).get(unit, [])
        results = {}
        for e in entries:
            frame = e.get('frame')
            if not frame:
                continue
            if e.get('form') != form_filter:
                continue
            if not _re.match(frame_pattern, frame):
                continue
            val = e.get('val')
            if val is not None:
                # Extract period key from frame
                if _re.match(r'^CY\d{4}$', frame):
                    period_key = frame[2:]  # '2024'
                elif _re.match(r'^CY\d{4}Q\dI?$', frame):
                    year = frame[2:6]
                    q = frame[7] if len(frame) > 7 else frame[6]
                    period_key = f'{year}-Q{q}'
                else:
                    continue
                results[period_key] = round(val / 1e9, 3)
        if results:
            return results
    return {}


def _build_sec_income_records(facts, form_filter, frame_pattern, logger):
    """Build income statement records from SEC EDGAR XBRL data."""
    field_values = {}
    for field, tags in _XBRL_INCOME_TAGS.items():
        if field == 'EBITDA':
            continue
        field_values[field] = _extract_xbrl_values(facts, tags, form_filter, frame_pattern)

    # Compute EBITDA = Operating Income + D&A (approximation)
    oi_vals = field_values.get('Operating Income', {})
    da_vals = _extract_xbrl_values(facts, _XBRL_DA_TAGS, form_filter, frame_pattern)
    ebitda_vals = {}
    for period in set(list(oi_vals.keys()) + list(da_vals.keys())):
        oi = oi_vals.get(period)
        da = da_vals.get(period)
        if oi is not None and da is not None:
            ebitda_vals[period] = round(oi + abs(da), 3)
    field_values['EBITDA'] = ebitda_vals

    # Collect all periods
    all_periods = set()
    for fv in field_values.values():
        all_periods.update(fv.keys())
    if not all_periods:
        return []

    sorted_periods = sorted(all_periods)
    records = []
    for i, period in enumerate(sorted_periods):
        rec = {'period': period}
        for field in ['Total Revenue', 'Cost Of Revenue', 'Gross Profit', 'Operating Income', 'EBITDA', 'Net Income']:
            rec[field] = field_values.get(field, {}).get(period)

        # Skip periods with no revenue
        if rec.get('Total Revenue') is None or rec['Total Revenue'] == 0:
            continue

        rev = rec['Total Revenue']
        if rev and rev > 0:
            gp = rec.get('Gross Profit')
            oi = rec.get('Operating Income')
            ni = rec.get('Net Income')
            eb = rec.get('EBITDA')
            rec['gross_margin'] = round(gp / rev * 100, 1) if gp is not None else None
            rec['operating_margin'] = round(oi / rev * 100, 1) if oi is not None else None
            rec['net_margin'] = round(ni / rev * 100, 1) if ni is not None else None
            rec['ebitda_margin'] = round(eb / rev * 100, 1) if eb is not None else None

        if i > 0:
            prev_period = sorted_periods[i - 1]
            prev_rev = field_values.get('Total Revenue', {}).get(prev_period)
            if prev_rev and prev_rev > 0 and rev:
                rec['revenue_growth_yoy'] = round((rev - prev_rev) / prev_rev * 100, 1)
            else:
                rec['revenue_growth_yoy'] = None
        else:
            rec['revenue_growth_yoy'] = None

        # Add 'year' key for annual records
        if '-' not in period:
            rec['year'] = period

        records.append(rec)
    return records


def _build_sec_balance_records(facts, form_filter, frame_pattern, logger):
    """Build balance sheet records from SEC EDGAR XBRL data."""
    field_values = {}
    for field, tags in _XBRL_BALANCE_TAGS.items():
        field_values[field] = _extract_xbrl_values(facts, tags, form_filter, frame_pattern)

    all_periods = set()
    for fv in field_values.values():
        all_periods.update(fv.keys())
    if not all_periods:
        return []

    sorted_periods = sorted(all_periods)
    records = []
    for period in sorted_periods:
        rec = {'period': period}
        for field in _XBRL_BALANCE_TAGS:
            rec[field] = field_values.get(field, {}).get(period)

        # Current ratio
        ca = rec.get('Current Assets')
        cl = rec.get('Current Liabilities')
        rec['Current Ratio'] = round(ca / cl, 2) if ca is not None and cl is not None and cl > 0 else None

        # Debt to equity
        td = rec.get('Total Debt')
        eq = rec.get('Stockholders Equity')
        rec['Debt to Equity'] = round(td / eq * 100, 1) if td is not None and eq is not None and eq > 0 else None

        records.append(rec)
    return records


def _build_sec_cashflow_records(facts, form_filter, frame_pattern, logger):
    """Build cash flow records from SEC EDGAR XBRL data."""
    field_values = {}
    for field, tags in _XBRL_CASHFLOW_TAGS.items():
        field_values[field] = _extract_xbrl_values(facts, tags, form_filter, frame_pattern)

    # Compute Free Cash Flow = Operating CF - |CapEx|
    ocf_vals = field_values.get('Operating Cash Flow', {})
    capex_vals = field_values.get('Capital Expenditure', {})
    fcf_vals = {}
    for period in set(list(ocf_vals.keys()) + list(capex_vals.keys())):
        ocf = ocf_vals.get(period)
        capex = capex_vals.get(period)
        if ocf is not None and capex is not None:
            fcf_vals[period] = round(ocf - abs(capex), 3)
    field_values['Free Cash Flow'] = fcf_vals

    all_periods = set()
    for fv in field_values.values():
        all_periods.update(fv.keys())
    if not all_periods:
        return []

    sorted_periods = sorted(all_periods)
    records = []
    for period in sorted_periods:
        rec = {'period': period}
        rec['Operating Cash Flow'] = field_values.get('Operating Cash Flow', {}).get(period)
        rec['Capital Expenditure'] = field_values.get('Capital Expenditure', {}).get(period)
        rec['Free Cash Flow'] = field_values.get('Free Cash Flow', {}).get(period)
        rec['Depreciation & Amortization'] = field_values.get('Depreciation & Amortization', {}).get(period)
        records.append(rec)
    return records


def _merge_sec_into_yfinance(yf_data, sec_data, logger, ticker):
    """Merge SEC EDGAR data into yfinance data, filling gaps only.
    Does NOT overwrite existing yfinance periods that have valid revenue data.
    """
    STATEMENT_KEYS = ['annual', 'quarterly', 'balance_sheet_annual', 'balance_sheet_quarterly',
                      'cashflow_annual', 'cashflow_quarterly']

    for key in STATEMENT_KEYS:
        yf_records = yf_data.get(key, [])
        sec_records = sec_data.get(key, [])
        if not sec_records:
            continue

        added = []

        # Step 1: Replace yfinance records with bad data ($0 revenue) in-place
        if key in ('annual', 'quarterly'):
            for i, rec in enumerate(yf_records):
                p = rec.get('period')
                rev = rec.get('Total Revenue')
                if p and (rev is None or rev == 0):
                    sec_match = next((s for s in sec_records if s.get('period') == p), None)
                    if sec_match and sec_match.get('Total Revenue') and sec_match['Total Revenue'] > 0:
                        yf_records[i] = sec_match
                        added.append(f'{p}(fixed)')

        # Step 2: Build set of periods now present with valid data
        existing_periods = set()
        for rec in yf_records:
            p = rec.get('period')
            if p:
                existing_periods.add(p)

        # Step 3: Add SEC records for truly missing periods
        for sec_rec in sec_records:
            p = sec_rec.get('period')
            if p and p not in existing_periods:
                yf_records.append(sec_rec)
                added.append(p)

        if added:
            logger.info(f'      SEC EDGAR filled {key}: {", ".join(added)}')

        # Re-sort and trim to last 5
        yf_records.sort(key=lambda r: r.get('period', ''))
        yf_data[key] = yf_records[-5:]


def _fetch_sec_fallback_for_ticker(ticker, cik, logger):
    """Fetch all financial statements from SEC EDGAR for one ticker.
    Returns a dict in the same shape as yfinance ticker_data.
    """
    facts = _fetch_sec_companyfacts(cik, logger)
    if not facts:
        return None

    ticker_data = {}

    # Annual income (duration items, frame like CY2024)
    ticker_data['annual'] = _build_sec_income_records(facts, '10-K', r'^CY\d{4}$', logger)
    # Quarterly income (duration items, frame like CY2024Q3)
    ticker_data['quarterly'] = _build_sec_income_records(facts, '10-Q', r'^CY\d{4}Q\d$', logger)

    # Annual balance sheet (instant items, frame like CY2024Q4I for year-end)
    ticker_data['balance_sheet_annual'] = _build_sec_balance_records(facts, '10-K', r'^CY\d{4}Q4I$', logger)
    # Remap balance_sheet_annual period keys from "2024-Q4" to "2024" to match yfinance convention
    for rec in ticker_data['balance_sheet_annual']:
        p = rec.get('period', '')
        if p.endswith('-Q4'):
            rec['period'] = p[:-3]
    # Quarterly balance sheet (instant items, frame like CY2024Q3I)
    ticker_data['balance_sheet_quarterly'] = _build_sec_balance_records(facts, '10-Q', r'^CY\d{4}Q\dI$', logger)

    # Annual cash flow (duration items, frame like CY2024)
    ticker_data['cashflow_annual'] = _build_sec_cashflow_records(facts, '10-K', r'^CY\d{4}$', logger)
    # Quarterly cash flow (duration items, frame like CY2024Q3)
    ticker_data['cashflow_quarterly'] = _build_sec_cashflow_records(facts, '10-Q', r'^CY\d{4}Q\d$', logger)

    return ticker_data


def fetch_historical_financials(market_data, logger):
    """Fetch multi-year financials: annual + quarterly income, balance sheet, and cash flow.
    Uses yfinance as primary source, then fills gaps from SEC EDGAR XBRL data.
    """
    logger.info('Fetching historical financials...')
    history = {}

    for ticker in market_data:
        try:
            stock = yf.Ticker(ticker)
            ticker_data = {}

            # Annual income statement
            fin = stock.financials
            if fin is not None and not fin.empty:
                annual_records = _build_income_records(fin, period_fmt='year')
                # Add 'year' key for backward compat with Equities/Peer dashboards
                for rec in annual_records:
                    rec['year'] = rec['period']
                ticker_data['annual'] = annual_records[-5:]  # last 5 years
            else:
                ticker_data['annual'] = []

            # Quarterly income statement
            qfin = stock.quarterly_financials
            if qfin is not None and not qfin.empty:
                ticker_data['quarterly'] = _build_income_records(qfin, period_fmt='quarter')[-5:]
            else:
                ticker_data['quarterly'] = []

            # Annual balance sheet
            bs = stock.balance_sheet
            if bs is not None and not bs.empty:
                ticker_data['balance_sheet_annual'] = _build_balance_sheet_records(bs, period_fmt='year')[-5:]
            else:
                ticker_data['balance_sheet_annual'] = []

            # Quarterly balance sheet
            qbs = stock.quarterly_balance_sheet
            if qbs is not None and not qbs.empty:
                ticker_data['balance_sheet_quarterly'] = _build_balance_sheet_records(qbs, period_fmt='quarter')[-5:]
            else:
                ticker_data['balance_sheet_quarterly'] = []

            # Annual cash flow
            cfn = stock.cashflow
            if cfn is not None and not cfn.empty:
                ticker_data['cashflow_annual'] = _build_cashflow_records(cfn, period_fmt='year')[-5:]
            else:
                ticker_data['cashflow_annual'] = []

            # Quarterly cash flow
            qcf = stock.quarterly_cashflow
            if qcf is not None and not qcf.empty:
                ticker_data['cashflow_quarterly'] = _build_cashflow_records(qcf, period_fmt='quarter')[-5:]
            else:
                ticker_data['cashflow_quarterly'] = []

            history[ticker] = ticker_data
            n_a = len(ticker_data['annual'])
            n_q = len(ticker_data['quarterly'])
            n_bsa = len(ticker_data['balance_sheet_annual'])
            n_bsq = len(ticker_data['balance_sheet_quarterly'])
            n_cfa = len(ticker_data['cashflow_annual'])
            n_cfq = len(ticker_data['cashflow_quarterly'])
            logger.info(f'    {ticker}: {n_a}yr/{n_q}qtr income, {n_bsa}yr/{n_bsq}qtr balance, {n_cfa}yr/{n_cfq}qtr cashflow')

        except Exception as e:
            logger.warning(f'    {ticker}: historical financials failed — {e}')

    logger.info(f'Historical financials: {len(history)}/{len(market_data)} tickers')

    # ── SEC EDGAR XBRL fallback: fill gaps where yfinance is missing data ──
    cik_map = _load_sec_cik_map()
    tickers_needing_sec = []
    for ticker, td in history.items():
        if ticker not in cik_map:
            continue
        annual = td.get('annual', [])
        # Check for missing recent data or corrupted $0 revenue
        has_gap = len(annual) < 3
        has_bad_data = any(r.get('Total Revenue') in (None, 0, 0.0) for r in annual)
        # Check if this ticker is behind other tickers' latest annual period
        all_annual_periods = set()
        for other_td in history.values():
            for r in other_td.get('annual', []):
                p = r.get('period')
                if p:
                    all_annual_periods.add(p)
        latest_global = max(all_annual_periods) if all_annual_periods else None
        latest_this = max((r['period'] for r in annual), default=None)
        is_behind = latest_global and latest_this and latest_this < latest_global

        if has_gap or has_bad_data or is_behind:
            tickers_needing_sec.append(ticker)

    if tickers_needing_sec:
        import time as _time
        logger.info(f'SEC EDGAR fallback for: {", ".join(tickers_needing_sec)}')
        for ticker in tickers_needing_sec:
            cik = cik_map[ticker]
            logger.info(f'    {ticker} (CIK {cik}): fetching from SEC EDGAR...')
            sec_data = _fetch_sec_fallback_for_ticker(ticker, cik, logger)
            if sec_data:
                _merge_sec_into_yfinance(history[ticker], sec_data, logger, ticker)
                td = history[ticker]
                n_a = len(td.get('annual', []))
                n_q = len(td.get('quarterly', []))
                logger.info(f'    {ticker} after merge: {n_a}yr/{n_q}qtr income')
            _time.sleep(0.15)  # Respect SEC rate limit (10 req/sec)

    return history


# =============================================================================
# PEER RANKINGS
# =============================================================================

def compute_peer_rankings(market_data, logger):
    """Compute within-category peer rankings and composite scores."""
    logger.info('Computing peer rankings...')

    # Metrics to rank (field, higher_is_better)
    rank_metrics = [
        ('market_cap_b', True),
        ('pe_trailing', False),       # lower P/E = cheaper
        ('revenue_growth', True),
        ('operating_margins', True),
        ('change_1y', True),
        ('ev_ebitda', False),         # lower EV/EBITDA = cheaper
    ]

    for category in CATEGORY_ORDER:
        peers = {t: d for t, d in market_data.items() if d['category'] == category}
        if not peers:
            continue

        for metric, higher_is_better in rank_metrics:
            # Get valid values
            # Filter to numeric values only — delisted tickers can leave string
            # sentinels (e.g. 'N/A') that break sort. Lesson #45.
            valid = [(t, d.get(metric)) for t, d in peers.items()
                     if isinstance(d.get(metric), (int, float)) and not isinstance(d.get(metric), bool)]
            if not valid:
                for t in peers:
                    market_data[t][f'rank_{metric}'] = None
                    market_data[t][f'pctile_{metric}'] = None
                continue

            # Sort and rank
            valid.sort(key=lambda x: x[1], reverse=higher_is_better)
            for rank, (t, val) in enumerate(valid, 1):
                market_data[t][f'rank_{metric}'] = rank
                # Percentile: 100 = best in group
                n = len(valid)
                market_data[t][f'pctile_{metric}'] = round((n - rank) / max(n - 1, 1) * 100, 1) if n > 1 else 50.0

            # Mark tickers with None values
            valid_tickers = {t for t, _ in valid}
            for t in peers:
                if t not in valid_tickers:
                    market_data[t][f'rank_{metric}'] = None
                    market_data[t][f'pctile_{metric}'] = None

        # Composite score (average of available percentiles)
        for t in peers:
            pctiles = [market_data[t].get(f'pctile_{m}') for m, _ in rank_metrics
                       if market_data[t].get(f'pctile_{m}') is not None]
            market_data[t]['composite_score'] = round(sum(pctiles) / len(pctiles), 1) if pctiles else None

    return market_data


# =============================================================================
# PRICE HISTORY & BIG MOVERS
# =============================================================================

def update_price_history(market_data, logger):
    """Update rolling 30-day price history."""
    logger.info('Updating price history...')
    history = {}
    if PRICE_HISTORY_FILE.exists():
        try:
            with open(PRICE_HISTORY_FILE) as f:
                history = json.load(f)
        except Exception:
            history = {}

    today = datetime.now().strftime('%Y-%m-%d')
    today_prices = {}
    today_mcaps = {}
    for ticker, data in market_data.items():
        if not isinstance(data, dict):
            continue
        if data.get('price_usd'):
            today_prices[ticker] = round(float(data['price_usd']), 2)
        if data.get('market_cap_b'):
            today_mcaps[ticker] = round(float(data['market_cap_b']), 3)

    history[today] = {
        'prices': today_prices,
        'market_caps': today_mcaps,
        'total_mcap': round(sum(today_mcaps.values()), 2),
    }

    # Prune to HISTORY_DAYS
    cutoff = (datetime.now() - timedelta(days=HISTORY_DAYS)).strftime('%Y-%m-%d')
    history = {k: v for k, v in history.items() if k >= cutoff}

    logger.info(f'  Price history: {len(history)} days tracked')
    return history


def compute_basket_index(price_history, market_data, label, basket_tickers, logger):
    """Compute market-cap-weighted index for a ticker basket, normalized to base 1000, plus benchmarks.

    `basket_tickers` is the explicit ticker subset to include (Inspection-11 or Flow & MOS-11).
    """
    logger.info(f'Computing {label} Index...')

    dates = sorted(price_history.keys())
    if len(dates) < 2:
        logger.warning(f'  Not enough price history for {label} index')
        return None

    # Base date = earliest date in history
    base_date = dates[0]
    base_prices = price_history[base_date].get('prices', {})
    base_mcaps = price_history[base_date].get('market_caps', {})

    # Use latest market caps as weights — restricted to the basket
    latest_mcaps_all = price_history[dates[-1]].get('market_caps', {})
    latest_mcaps = {t: mc for t, mc in latest_mcaps_all.items() if t in basket_tickers}
    total_mcap = sum(latest_mcaps.values())
    if total_mcap == 0:
        logger.warning(f'  Zero total market cap for {label} basket — skipping')
        return None
    weights = {t: mc / total_mcap for t, mc in latest_mcaps.items()}

    # Per-ticker base price = earliest non-zero price in price_history.
    # Tickers added to the universe AFTER the price_history base_date have no entry in base_prices,
    # so we walk forward to find each ticker's first observed price and use that as its base.
    per_ticker_base = {}
    for ticker in weights.keys():
        bp = base_prices.get(ticker)
        if bp and bp > 0:
            per_ticker_base[ticker] = bp
            continue
        for d in dates:
            p = price_history[d].get('prices', {}).get(ticker)
            if p and p > 0:
                per_ticker_base[ticker] = p
                break

    # Compute index for each date (with forward-fill for missing tickers)
    index_series = []
    last_known_prices = dict(base_prices)  # forward-fill tracker
    for date in dates:
        day_prices = price_history[date].get('prices', {})
        if not day_prices:
            continue
        # Update forward-fill tracker with today's prices
        for t, p in day_prices.items():
            if p and p > 0:
                last_known_prices[t] = p
        weighted_return = 0
        weight_sum = 0
        for ticker, weight in weights.items():
            price = day_prices.get(ticker) or last_known_prices.get(ticker)
            base = per_ticker_base.get(ticker)
            if price and base and base > 0:
                ret = price / base
                weighted_return += ret * weight
                weight_sum += weight
        if weight_sum > 0:
            index_val = round(1000 * weighted_return / weight_sum, 2)
            index_series.append({'date': date, 'value': index_val})

    if not index_series:
        logger.warning(f'  No valid base prices for {label} basket tickers — skipping')
        return None

    # Fetch benchmarks
    benchmarks = {}
    benchmark_tickers = {'^GSPC': 'S&P 500', '^RUT': 'Russell 2000', 'XLI': 'Industrials ETF'}
    for bench_ticker, bench_name in benchmark_tickers.items():
        try:
            import yfinance as yf
            stock = yf.Ticker(bench_ticker)
            hist = stock.history(start=base_date, end=datetime.now().strftime('%Y-%m-%d'))
            if hist.empty:
                continue
            base_val = hist['Close'].iloc[0]
            bench_series = []
            for idx, row in hist.iterrows():
                d = idx.strftime('%Y-%m-%d')
                bench_series.append({'date': d, 'value': round(1000 * row['Close'] / base_val, 2)})
            benchmarks[bench_name] = bench_series
            logger.info(f'  Benchmark {bench_name}: {len(bench_series)} days')
        except Exception as e:
            logger.warning(f'  Failed to fetch benchmark {bench_ticker}: {e}')

    # Current values
    current = index_series[-1]['value']
    prev = index_series[-2]['value'] if len(index_series) > 1 else current
    daily_change = round(current - prev, 2)
    daily_pct = round((current - prev) / prev * 100, 2) if prev else 0

    # YTD
    ytd_start = None
    year_start = datetime.now().strftime('%Y') + '-01-01'
    for pt in index_series:
        if pt['date'] >= year_start:
            ytd_start = pt['value']
            break
    ytd_pct = round((current - ytd_start) / ytd_start * 100, 2) if ytd_start and ytd_start > 0 else None

    result = {
        'index': index_series,
        'benchmarks': benchmarks,
        'weights': {t: round(w*100, 1) for t, w in weights.items()},
        'current': current,
        'daily_change': daily_change,
        'daily_change_pct': daily_pct,
        'ytd_pct': ytd_pct,
        'base_date': base_date,
        'generated_at': datetime.now().isoformat(),
    }

    ytd_str = f'{ytd_pct:+.2f}%' if ytd_pct is not None else 'n/a'
    logger.info(f'  {label} Index: {current:.2f} ({daily_pct:+.2f}% today, YTD: {ytd_str})')
    return result


def detect_big_movers(market_data, price_history, logger):
    """Detect big movers vs previous day in history."""
    logger.info('Detecting big movers...')
    big_movers = []
    hist_dates = sorted(price_history.keys())

    if len(hist_dates) >= 2:
        prev_date = hist_dates[-2]
        prev_prices = price_history[prev_date].get('prices', {})
        for ticker, data in market_data.items():
            curr = data.get('price_usd')
            prev = prev_prices.get(ticker)
            if curr and prev and prev > 0:
                pct = ((curr - prev) / prev) * 100
                if abs(pct) >= BIG_MOVER_THRESHOLD:
                    big_movers.append({
                        'ticker': ticker,
                        'display_ticker': data.get('display_ticker', ticker),
                        'company': data.get('name', ''),
                        'category': data.get('category', ''),
                        'prev_price': prev,
                        'curr_price': curr,
                        'pct_change': round(pct, 2),
                        'market_cap_b': data.get('market_cap_b'),
                        'prev_date': prev_date,
                    })
    else:
        # First run fallback
        for ticker, data in market_data.items():
            chg = data.get('daily_change_pct')
            if chg is not None:
                pct = float(chg)
                if abs(pct) < 1:
                    pct *= 100
                if abs(pct) >= BIG_MOVER_THRESHOLD:
                    big_movers.append({
                        'ticker': ticker,
                        'display_ticker': data.get('display_ticker', ticker),
                        'company': data.get('name', ''),
                        'category': data.get('category', ''),
                        'prev_price': None,
                        'curr_price': data.get('price_usd'),
                        'pct_change': round(pct, 2),
                        'market_cap_b': data.get('market_cap_b'),
                        'prev_date': 'previous close',
                    })

    big_movers.sort(key=lambda x: abs(x['pct_change']), reverse=True)
    logger.info(f'  Found {len(big_movers)} big movers (>{BIG_MOVER_THRESHOLD}%)')
    return big_movers


# =============================================================================
# FRED / CUSTOMER INDUSTRY DATA
# =============================================================================

def fetch_fred_series(series_id, days=90, logger=None, retries=2):
    """Fetch a FRED series as CSV (free, no API key). Retries on timeout."""
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    url = f'https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}&cosd={start_date}&coed={end_date}'
    for attempt in range(retries + 1):
        try:
            req = Request(url, headers={'User-Agent': USER_AGENT})
            with urlopen(req, timeout=30) as resp:
                text = resp.read().decode('utf-8')
            rows = []
            reader = csv.DictReader(text.strip().split('\n'))
            for row in reader:
                val = row.get(series_id, '').strip()
                date = (row.get('observation_date') or row.get('DATE') or '').strip()
                if val and val != '.' and date:
                    try:
                        rows.append({'date': date, 'value': float(val)})
                    except ValueError:
                        pass
            return rows
        except Exception as e:
            if attempt < retries:
                import time
                time.sleep(2)  # brief pause before retry
                continue
            if logger:
                logger.warning(f'  Failed to fetch FRED {series_id} after {retries+1} attempts: {e}')
            return []


def fetch_industry_indicators(logger):
    """Fetch O&G and manufacturing indicators from FRED."""
    logger.info('Fetching customer industry indicators from FRED...')
    indicators = {
        'fetched_at': datetime.now().isoformat(),
        'oil_gas': {},
        'manufacturing': {},
        'customer_industry_weights': CUSTOMER_INDUSTRY_WEIGHTS,
        'fred_history': {},
    }

    # Load previous indicators as fallback for failed fetches
    prev_indicators = {}
    if INDUSTRY_INDICATORS_FILE.exists():
        try:
            with open(INDUSTRY_INDICATORS_FILE) as pf:
                prev_indicators = json.load(pf)
        except Exception:
            pass

    for series_id, label in FRED_SERIES.items():
        rows = fetch_fred_series(series_id, days=365, logger=logger)
        if not rows:
            # Try to preserve previous data for this series
            key = series_id.lower()
            for cat in ['oil_gas', 'drilling', 'aerospace', 'power', 'infrastructure', 'manufacturing']:
                if key in prev_indicators.get(cat, {}):
                    indicators.setdefault(cat, {})[key] = prev_indicators[cat][key]
                    logger.warning(f'  No fresh data for {series_id} ({label}) — using previous value')
                    break
                elif key in prev_indicators.get('fred_history', {}):
                    indicators['fred_history'][key] = prev_indicators['fred_history'][key]
                    break
            else:
                logger.warning(f'  No data for {series_id} ({label})')
            continue

        current = rows[-1]['value']
        current_date = rows[-1]['date']

        # 30-day change
        cutoff_30 = (datetime.now() - timedelta(days=35)).strftime('%Y-%m-%d')
        recent_30 = [r for r in rows if r['date'] <= cutoff_30]
        change_30d = None
        if recent_30:
            old_val = recent_30[-1]['value']
            if old_val > 0:
                change_30d = round((current - old_val) / old_val * 100, 1)

        # 90-day change
        cutoff_90 = (datetime.now() - timedelta(days=95)).strftime('%Y-%m-%d')
        recent_90 = [r for r in rows if r['date'] <= cutoff_90]
        change_90d = None
        if recent_90:
            old_val = recent_90[-1]['value']
            if old_val > 0:
                change_90d = round((current - old_val) / old_val * 100, 1)

        key = series_id.lower()
        entry = {
            'current': current,
            'date': current_date,
            'change_30d_pct': change_30d,
            'change_90d_pct': change_90d,
            'label': label,
        }

        # Categorize into groups
        OG_SERIES = {'DCOILWTICO', 'DCOILBRENTEU', 'DHHNGSP', 'DHOILNYH', 'GASREGW'}
        INFRA_SERIES = {'TLRESCONS', 'TLNRESCONS', 'TLPWRCONS', 'PBHWYCONS', 'TLMFGCONS', 'DGORDER'}
        DRILLING_SERIES = {'IPN213111N', 'CAPUTLG211S', 'CAPUTLG324S', 'IPG211S'}
        AERO_SERIES = {'ADEFNO', 'UNAPNO'}
        POWER_SERIES = {'IPN221113S', 'CAPUTLG2211S'}

        if series_id in OG_SERIES:
            indicators['oil_gas'][key] = entry
        elif series_id in INFRA_SERIES:
            if 'infrastructure' not in indicators:
                indicators['infrastructure'] = {}
            indicators['infrastructure'][key] = entry
        elif series_id in DRILLING_SERIES:
            if 'drilling' not in indicators:
                indicators['drilling'] = {}
            indicators['drilling'][key] = entry
        elif series_id in AERO_SERIES:
            if 'aerospace' not in indicators:
                indicators['aerospace'] = {}
            indicators['aerospace'][key] = entry
        elif series_id in POWER_SERIES:
            if 'power' not in indicators:
                indicators['power'] = {}
            indicators['power'][key] = entry
        else:
            indicators['manufacturing'][key] = entry

        # Store last 90 days for charts
        last_90 = rows[-90:] if len(rows) >= 90 else rows
        indicators['fred_history'][key] = last_90

        logger.info(f'  {label}: {current:.2f} (30d: {change_30d:+.1f}%)' if change_30d else f'  {label}: {current:.2f}')

    return indicators


# =============================================================================
# NEWS & OSHA DATA
# =============================================================================

def load_news_csv(n=25, logger=None):
    """Load most recent N news articles from both industry and company news CSVs."""
    all_rows = []

    # Load industry news
    industry_csv = INDUSTRY_DIR / 'News' / 'industry_news.csv'
    if industry_csv.exists():
        try:
            with open(industry_csv, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row['type'] = 'industry'
                    all_rows.append(row)
        except Exception as e:
            if logger:
                logger.warning(f'  Failed to load industry news CSV: {e}')
    elif logger:
        logger.warning('  Industry news CSV not found')

    # Load company news
    company_csv = INDUSTRY_DIR / 'News' / 'company_news.csv'
    if company_csv.exists():
        try:
            with open(company_csv, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row['type'] = 'company'
                    all_rows.append(row)
        except Exception as e:
            if logger:
                logger.warning(f'  Failed to load company news CSV: {e}')

    all_rows.sort(key=lambda x: x.get('date', ''), reverse=True)
    return all_rows[:n]


def load_osha_summary(logger=None):
    """Summarize OSHA inspection data."""
    csv_path = INDUSTRY_DIR / 'OSHA' / 'osha_inspections.csv'
    if not csv_path.exists():
        if logger:
            logger.warning('  OSHA CSV not found')
        return {'total': 0, 'with_violations': 0, 'recent_90d': 0, 'by_company': []}
    try:
        with open(csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        total = len(rows)
        with_violations = sum(1 for r in rows if int(r.get('violations_count', 0) or 0) > 0)

        cutoff_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        recent_90d = sum(1 for r in rows if r.get('date_opened', '') >= cutoff_90)

        # Group by establishment
        companies = {}
        for r in rows:
            name = r.get('establishment_name', 'Unknown')
            if name not in companies:
                companies[name] = {'name': name, 'count': 0, 'violations': 0, 'states': set(), 'latest': ''}
            companies[name]['count'] += 1
            companies[name]['violations'] += int(r.get('violations_count', 0) or 0)
            companies[name]['states'].add(r.get('state', ''))
            d = r.get('date_opened', '')
            if d > companies[name]['latest']:
                companies[name]['latest'] = d

        by_company = sorted(companies.values(), key=lambda x: x['count'], reverse=True)
        for c in by_company:
            c['states'] = list(c['states'])

        return {'total': total, 'with_violations': with_violations, 'recent_90d': recent_90d, 'by_company': by_company}
    except Exception as e:
        if logger:
            logger.warning(f'  Failed to load OSHA CSV: {e}')
        return {'total': 0, 'with_violations': 0, 'recent_90d': 0, 'by_company': []}


# =============================================================================
# EXCEL GENERATION
# =============================================================================

HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Arial', size=10)
BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
GREEN_FILL = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
RED_FILL = PatternFill(start_color='4a1a1a', end_color='4a1a1a', fill_type='solid')


def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'


def write_row(ws, row_num, values, formats=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row_num, col, val)
        cell.border = BORDER
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if formats and col - 1 < len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]


def generate_excel(market_data, price_history, big_movers, industry_indicators, logger):
    """Generate complete Excel workbook."""
    logger.info('Generating Excel workbook...')
    FINANCIALS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    tickers_sorted = sorted(market_data.keys(), key=lambda t: market_data[t].get('market_cap_b', 0), reverse=True)

    # ── Sheet 1: Market Overview ──
    ws = wb.active
    ws.title = 'Market Overview'
    headers = ['Ticker', 'Company', 'Category', 'Price ($)', 'Daily Chg %',
               '52W High', '52W Low', '% Off High', 'Market Cap ($B)', 'EV ($B)',
               'P/E', 'Fwd P/E', 'EV/EBITDA', 'Beta', '1Y Chg %', 'YTD Chg %']
    write_header(ws, headers)
    for i, t in enumerate(tickers_sorted, 2):
        d = market_data[t]
        vals = [d['display_ticker'], d['name'], d['category'],
                d['price_usd'], d.get('daily_change_pct'),
                d.get('fifty_two_week_high'), d.get('fifty_two_week_low'),
                d.get('pct_off_high', 0) / 100 if d.get('pct_off_high') is not None else None,
                d['market_cap_b'], fmt_val(d.get('enterprise_value_b')),
                fmt_val(d.get('pe_trailing')), fmt_val(d.get('pe_forward')),
                fmt_val(d.get('ev_ebitda')), fmt_val(d.get('beta')),
                d.get('change_1y', 0) / 100, d.get('change_ytd', 0) / 100]
        fmts = [None, None, None, '$#,##0.00', '0.0%',
                '$#,##0.00', '$#,##0.00', '0.0%', '#,##0.000', '#,##0.000',
                '0.0', '0.0', '0.0', '0.00', '0.0%', '0.0%']
        write_row(ws, i, vals, fmts)
    widths = [8, 22, 20, 10, 10, 10, 10, 10, 12, 12, 8, 8, 10, 8, 10, 10]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 2: Valuation ──
    ws2 = wb.create_sheet('Valuation')
    headers = ['Ticker', 'Company', 'Category', 'Price', 'MCap ($B)',
               'P/E (TTM)', 'P/E (Fwd)', 'EV/EBITDA', 'EV/Revenue', 'P/B', 'P/S',
               'EPS (TTM)', 'EPS (Fwd)']
    write_header(ws2, headers)
    for i, t in enumerate(tickers_sorted, 2):
        d = market_data[t]
        vals = [d['display_ticker'], d['name'], d['category'], d['price_usd'], d['market_cap_b'],
                fmt_val(d.get('pe_trailing')), fmt_val(d.get('pe_forward')),
                fmt_val(d.get('ev_ebitda')), fmt_val(d.get('ev_revenue')),
                fmt_val(d.get('price_to_book')), fmt_val(d.get('price_to_sales')),
                fmt_val(d.get('eps_trailing')), fmt_val(d.get('eps_forward'))]
        fmts = [None, None, None, '$#,##0.00', '#,##0.000',
                '0.0', '0.0', '0.0', '0.0', '0.0', '0.0', '$#,##0.00', '$#,##0.00']
        write_row(ws2, i, vals, fmts)
    for j, w in enumerate([8, 22, 20, 10, 12, 8, 8, 10, 10, 8, 8, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 3: Financial Health ──
    ws3 = wb.create_sheet('Financial Health')
    headers = ['Ticker', 'Company', 'Category', 'Revenue ($B)', 'Rev Growth',
               'Gross Margin', 'Op Margin', 'Net Margin', 'EBITDA Margin',
               'ROE', 'ROA', 'EBITDA ($B)', 'FCF ($B)',
               'Total Debt ($B)', 'Total Cash ($B)', 'D/E', 'Current Ratio']
    write_header(ws3, headers)
    for i, t in enumerate(tickers_sorted, 2):
        d = market_data[t]
        vals = [d['display_ticker'], d['name'], d['category'],
                fmt_val(d.get('revenue_b')), fmt_val(d.get('revenue_growth')),
                fmt_val(d.get('gross_margins')), fmt_val(d.get('operating_margins')),
                fmt_val(d.get('profit_margins')), fmt_val(d.get('ebitda_margins')),
                fmt_val(d.get('roe')), fmt_val(d.get('roa')),
                fmt_val(d.get('ebitda_b')), fmt_val(d.get('free_cashflow_b')),
                fmt_val(d.get('total_debt_b')), fmt_val(d.get('total_cash_b')),
                fmt_val(d.get('debt_to_equity')), fmt_val(d.get('current_ratio'))]
        fmts = [None, None, None, '#,##0.000', '0.0%',
                '0.0%', '0.0%', '0.0%', '0.0%',
                '0.0%', '0.0%', '#,##0.000', '#,##0.000',
                '#,##0.000', '#,##0.000', '0.0', '0.0']
        write_row(ws3, i, vals, fmts)
    for j, w in enumerate([8, 22, 20, 12, 10, 12, 10, 10, 12, 8, 8, 12, 10, 12, 12, 8, 12], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 4: Analyst & Sentiment ──
    ws4 = wb.create_sheet('Analyst & Sentiment')
    headers = ['Ticker', 'Company', 'Price', 'Rec Key', '# Analysts',
               'Target Mean', 'Target High', 'Target Low', 'Upside %',
               'Short Ratio', 'Short % Float', '% Insiders', '% Institutions']
    write_header(ws4, headers)
    for i, t in enumerate(tickers_sorted, 2):
        d = market_data[t]
        vals = [d['display_ticker'], d['name'], d['price_usd'],
                fmt_val(d.get('recommendation_key')), fmt_val(d.get('num_analysts')),
                fmt_val(d.get('target_mean')), fmt_val(d.get('target_high')),
                fmt_val(d.get('target_low')),
                d.get('upside_pct', 0) / 100 if d.get('upside_pct') is not None else None,
                fmt_val(d.get('short_ratio')), fmt_val(d.get('short_pct_float')),
                fmt_val(d.get('held_pct_insiders')), fmt_val(d.get('held_pct_institutions'))]
        fmts = [None, None, '$#,##0.00', None, '#,##0',
                '$#,##0.00', '$#,##0.00', '$#,##0.00', '0.0%',
                '0.0', '0.0%', '0.0%', '0.0%']
        write_row(ws4, i, vals, fmts)
    for j, w in enumerate([8, 22, 10, 10, 10, 12, 12, 12, 10, 10, 12, 10, 12], 1):
        ws4.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 5: Technical ──
    ws5 = wb.create_sheet('Technical')
    headers = ['Ticker', 'Company', 'Price', '50-Day MA', '200-Day MA',
               'vs 50DMA %', 'vs 200DMA %', '52W High', '52W Low', '% Off High',
               '1Y Chg %', 'YTD Chg %', '1Y Trend', 'Beta']
    write_header(ws5, headers)
    for i, t in enumerate(tickers_sorted, 2):
        d = market_data[t]
        vals = [d['display_ticker'], d['name'], d['price_usd'],
                fmt_val(d.get('fifty_day_avg')), fmt_val(d.get('two_hundred_day_avg')),
                d.get('vs_50dma_pct', 0) / 100 if d.get('vs_50dma_pct') is not None else None,
                d.get('vs_200dma_pct', 0) / 100 if d.get('vs_200dma_pct') is not None else None,
                d.get('fifty_two_week_high'), d.get('fifty_two_week_low'),
                d.get('pct_off_high', 0) / 100 if d.get('pct_off_high') is not None else None,
                d.get('change_1y', 0) / 100, d.get('change_ytd', 0) / 100,
                fmt_val(d.get('trend_1y')), fmt_val(d.get('beta'))]
        fmts = [None, None, '$#,##0.00', '$#,##0.00', '$#,##0.00',
                '0.0%', '0.0%', '$#,##0.00', '$#,##0.00', '0.0%',
                '0.0%', '0.0%', '0.0', '0.00']
        write_row(ws5, i, vals, fmts)
    for j, w in enumerate([8, 22, 10, 10, 12, 10, 10, 10, 10, 10, 10, 10, 10, 8], 1):
        ws5.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 6: Big Movers ──
    ws6 = wb.create_sheet('Big Movers')
    headers = ['Ticker', 'Company', 'Category', 'Prev Price', 'Current Price',
               'Change %', 'MCap ($B)', 'vs Date']
    write_header(ws6, headers)
    for i, m in enumerate(big_movers, 2):
        vals = [m['display_ticker'], m['company'], m['category'],
                m.get('prev_price'), m['curr_price'],
                m['pct_change'] / 100, m.get('market_cap_b'), m.get('prev_date')]
        fmts = [None, None, None, '$#,##0.00', '$#,##0.00',
                '0.00%', '#,##0.000', None]
        write_row(ws6, i, vals, fmts)
        # Color code change
        cell = ws6.cell(i, 6)
        if m['pct_change'] > 0:
            cell.font = Font(name='Arial', size=10, bold=True, color='008000')
        else:
            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')
    if not big_movers:
        ws6.cell(2, 1, 'No big movers today').font = Font(italic=True, name='Arial', size=10, color='666666')
    for j, w in enumerate([8, 22, 20, 12, 12, 10, 12, 12], 1):
        ws6.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 7: Peer Comparison ──
    ws7 = wb.create_sheet('Peer Comparison')
    headers = ['Ticker', 'Company', 'Category', 'MCap ($B)', 'P/E',
               'Rev Growth', 'Op Margin', '1Y Return', 'EV/EBITDA', 'Composite Score']
    write_header(ws7, headers)
    row_idx = 2
    for category in CATEGORY_ORDER:
        peers = [(t, d) for t, d in market_data.items() if d['category'] == category]
        peers.sort(key=lambda x: x[1].get('composite_score') or 0, reverse=True)
        for t, d in peers:
            vals = [d['display_ticker'], d['name'], d['category'], d['market_cap_b'],
                    fmt_val(d.get('pe_trailing')),
                    fmt_val(d.get('revenue_growth')),
                    fmt_val(d.get('operating_margins')),
                    d.get('change_1y', 0) / 100,
                    fmt_val(d.get('ev_ebitda')),
                    fmt_val(d.get('composite_score'))]
            fmts = [None, None, None, '#,##0.000', '0.0',
                    '0.0%', '0.0%', '0.0%', '0.0', '0.0']
            write_row(ws7, row_idx, vals, fmts)
            row_idx += 1
        # Blank separator between categories
        row_idx += 1
    for j, w in enumerate([8, 22, 20, 12, 8, 10, 10, 10, 10, 14], 1):
        ws7.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 8: Summary Stats ──
    ws8 = wb.create_sheet('Summary Stats')
    total_mcap = sum(d['market_cap_b'] for d in market_data.values())
    pe_vals = [d['pe_trailing'] for d in market_data.values()
               if isinstance(d.get('pe_trailing'), (int, float))
               and not isinstance(d.get('pe_trailing'), bool)]
    ev_vals = [d['ev_ebitda'] for d in market_data.values()
               if isinstance(d.get('ev_ebitda'), (int, float))
               and not isinstance(d.get('ev_ebitda'), bool)]
    best = max(market_data.values(), key=lambda x: x.get('change_1y', 0)) if market_data else None
    worst = min(market_data.values(), key=lambda x: x.get('change_1y', 0)) if market_data else None

    stats = [
        ['Metric', 'Value'],
        ['Total Companies', len(market_data)],
        ['Total Market Cap ($B)', round(total_mcap, 1)],
        ['Median P/E (TTM)', round(statistics.median(pe_vals), 1) if pe_vals else 'N/A'],
        ['Median EV/EBITDA', round(statistics.median(ev_vals), 1) if ev_vals else 'N/A'],
        ['', ''],
        ['Best 1Y Performer', f'{best["name"]} ({best["change_1y"]:+.1f}%)' if best else 'N/A'],
        ['Worst 1Y Performer', f'{worst["name"]} ({worst["change_1y"]:+.1f}%)' if worst else 'N/A'],
        ['', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]

    # Per-category stats
    for cat in CATEGORY_ORDER:
        peers = [d for d in market_data.values() if d['category'] == cat]
        if not peers:
            continue
        cat_mcap = sum(d['market_cap_b'] for d in peers)
        cat_pe = [d['pe_trailing'] for d in peers if d.get('pe_trailing')]
        stats.append(['', ''])
        stats.append([f'--- {cat} ---', ''])
        stats.append([f'  Companies', len(peers)])
        stats.append([f'  Total MCap ($B)', round(cat_mcap, 1)])
        stats.append([f'  Median P/E', round(statistics.median(cat_pe), 1) if cat_pe else 'N/A'])

    for row_num, (label, val) in enumerate(stats, 1):
        ws8.cell(row_num, 1, label).font = Font(name='Arial', size=10, bold=('---' in str(label) or row_num == 1))
        cell = ws8.cell(row_num, 2, val)
        cell.font = NORMAL_FONT
        if isinstance(val, float):
            cell.number_format = '#,##0.0'
    ws8.column_dimensions['A'].width = 30
    ws8.column_dimensions['B'].width = 35

    # ── Sheet 9: Price History ──
    ws9 = wb.create_sheet('Price History')
    hist_dates = sorted(price_history.keys())
    ph_headers = ['Ticker'] + hist_dates
    write_header(ws9, ph_headers)
    for i, t in enumerate(sorted(market_data.keys()), 2):
        ws9.cell(i, 1, market_data[t]['display_ticker']).font = Font(name='Arial', size=10, bold=True)
        ws9.cell(i, 1).border = BORDER
        for j, d in enumerate(hist_dates, 2):
            p = price_history.get(d, {}).get('prices', {}).get(t)
            cell = ws9.cell(i, j, p)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            if p:
                cell.number_format = '$#,##0.00'
    ws9.column_dimensions['A'].width = 10
    for j in range(2, len(hist_dates) + 2):
        ws9.column_dimensions[get_column_letter(j)].width = 12

    wb.save(str(EXCEL_FILE))
    logger.info(f'Excel saved: {EXCEL_FILE}')


# =============================================================================
# HTML DATA INJECTION
# =============================================================================

def inject_json_into_html(html_path, var_name, json_data, logger):
    """Inject JSON data inline into an HTML file between marker comments."""
    if not os.path.exists(html_path):
        logger.info(f'  Skipping (not found): {html_path}')
        return

    with open(html_path, 'r') as f:
        content = f.read()

    pattern = r'(// INJECTED_DATA_START\s*\n)\s*var ' + re.escape(var_name) + r' = .*?;\s*\n(\s*// INJECTED_DATA_END)'
    json_str = json.dumps(json_data, indent=2, default=str)

    def replacer(m):
        return m.group(1) + '        var ' + var_name + ' = ' + json_str + ';\n        ' + m.group(2)

    new_content = re.sub(pattern, replacer, content, flags=re.DOTALL)

    if new_content != content:
        with open(html_path, 'w') as f:
            f.write(new_content)
        logger.info(f'  Injected {var_name} into {os.path.basename(html_path)}')
    else:
        logger.info(f'  No marker found for {var_name} in {os.path.basename(html_path)} — skipping')


# =============================================================================
# TRANSCRIPT SCANNING & HTML GENERATION
# =============================================================================

# Map transcript-file tickers to dashboard tickers
TRANSCRIPT_TICKER_MAP = {
    'BVRDF': 'BVI.PA',
    'ITRKY': 'ITRK.L',
    'ALQ':   'ALQ.AX',
    'ULSLF': 'ULS',
    # Direct matches: MG, TISI, XPRO, TRNS, THR, CLH, EXP, OSIS, ESE, NSSC, MHH, VPG, etc.
}

def _extract_summary(text, max_words=300):
    """Extract a summary from the CEO/CFO prepared remarks, skipping boilerplate."""
    lines = text.split('\n')
    summary_parts = []
    in_exec = False
    exec_speaker_count = 0
    word_count = 0

    # Phrases that indicate boilerplate / legal disclaimers to skip
    skip_phrases = [
        'forward-looking statement', 'safe harbor', 'non-gaap', 'non-u.s. gaap',
        'reconciliation of these', 'sec filings', 'form 10-k', 'form 8-k',
        'risks and uncertainties', 'actual results to differ',
        'i will now turn', 'let me turn the call', 'i\'ll turn the call',
        'welcome to', 'welcome everyone', 'good morning, everyone, and welcome',
        'conference operator', 'operator instructions',
    ]

    for line in lines:
        stripped = line.strip()

        # Detect Q&A section start
        if any(phrase in stripped.lower() for phrase in ['question-and-answer', 'q&a session', 'questions and answers',
                                                          'open the line for questions', 'open it up for questions',
                                                          'open the floor for questions', 'begin the question']):
            break

        # Detect speaker header
        if stripped.startswith('[') and stripped.endswith(']') and ' - ' in stripped:
            role = stripped.split(' - ')[-1].rstrip(']').strip().lower()
            if 'operator' in role:
                in_exec = False
                continue
            if 'executive' in role or 'ceo' in role or 'cfo' in role or 'president' in role:
                exec_speaker_count += 1
                in_exec = True
                continue
            else:
                if word_count > 50:
                    break
                in_exec = False
                continue

        if in_exec and stripped:
            # Skip boilerplate paragraphs (legal disclaimers, introductions)
            lower = stripped.lower()
            if any(bp in lower for bp in skip_phrases):
                continue
            # Skip very short handoff lines
            if len(stripped.split()) < 8 and any(w in lower for w in ['turn', 'over to', 'thank you']):
                continue

            words = stripped.split()
            if word_count + len(words) > max_words:
                remaining = max_words - word_count
                summary_parts.append(' '.join(words[:remaining]) + '...')
                word_count = max_words
                break
            summary_parts.append(stripped)
            word_count += len(words)

    return ' '.join(summary_parts) if summary_parts else ''


def _generate_transcript_html(transcript_data, output_path, company_name, period, date_str):
    """Generate a standalone HTML page for a single transcript."""
    title = f'{company_name} — {period} Earnings Call'
    # Process transcript text into formatted HTML
    lines = transcript_data.split('\n')
    body_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith('====='):
            body_lines.append('<hr style="border-color:rgba(255,255,255,0.1);margin:24px 0;">')
        elif stripped.startswith('[') and stripped.endswith(']') and ' - ' in stripped:
            speaker = stripped[1:-1]
            name_part, role_part = speaker.rsplit(' - ', 1)
            body_lines.append(f'<div class="speaker"><span class="speaker-name">{name_part}</span> <span class="speaker-role">— {role_part}</span></div>')
        elif stripped.startswith(('[Operator', 'Date:', 'Type:', 'Company:')):
            body_lines.append(f'<div class="meta-line">{stripped}</div>')
        else:
            body_lines.append(f'<p>{stripped}</p>')

    body_html = '\n'.join(body_lines)

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script src="../auth.js"></script>
    <style>
        :root {{
            --bg-primary: #0f1117;
            --bg-card: #1a1d29;
            --text-primary: #e8eaed;
            --text-secondary: #9aa0a6;
            --accent: #4fc3f7;
            --radius: 10px;
            --border: 1px solid rgba(255,255,255,0.06);
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: var(--bg-primary);
            color: var(--text-primary);
            line-height: 1.7;
        }}
        .container {{ max-width: 900px; margin: 0 auto; padding: 24px; }}
        .back-link {{
            display: inline-flex; align-items: center; gap: 6px;
            color: var(--accent); text-decoration: none; font-size: 13px;
            margin-bottom: 20px;
        }}
        .back-link:hover {{ text-decoration: underline; }}
        .header {{
            background: linear-gradient(135deg, #1a237e 0%, #0d47a1 50%, #01579b 100%);
            color: #fff; padding: 32px; border-radius: var(--radius);
            margin-bottom: 28px;
        }}
        .header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
        .header .date {{ font-size: 13px; opacity: 0.7; }}
        .transcript-body {{
            background: var(--bg-card);
            border: var(--border);
            border-radius: var(--radius);
            padding: 32px;
        }}
        .transcript-body p {{
            margin-bottom: 16px;
            font-size: 14px;
            color: var(--text-primary);
        }}
        .speaker {{
            margin-top: 28px;
            margin-bottom: 8px;
            padding: 8px 12px;
            background: rgba(79,195,247,0.08);
            border-radius: 6px;
            border-left: 3px solid var(--accent);
        }}
        .speaker-name {{ font-weight: 600; color: var(--accent); font-size: 14px; }}
        .speaker-role {{ color: var(--text-secondary); font-size: 12px; }}
        .meta-line {{ font-size: 13px; color: var(--text-secondary); margin-bottom: 4px; }}
        .footer {{
            text-align: center; padding: 20px; font-size: 12px;
            color: var(--text-secondary); border-top: var(--border); margin-top: 40px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <a class="back-link" href="../TIC_NDT_Company_Summary.html">← Back to Company Summary</a>
        <div class="header">
            <h1>{title}</h1>
            <div class="date">{date_str}</div>
        </div>
        <div class="transcript-body">
            {body_html}
        </div>
    </div>
    <div class="footer">Inspection Intel &mdash; Proprietary &amp; Confidential</div>
</body>
</html>'''
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        f.write(html)


def scan_and_generate_transcripts(logger):
    """Scan Companies for transcripts, generate HTML pages, return metadata."""
    research_dir = PROJECT_DIR / 'Companies'
    transcripts_out_dir = DASHBOARD_DIR / 'transcripts'
    transcripts_out_dir.mkdir(parents=True, exist_ok=True)

    transcript_manifest = {}  # dashboard_ticker -> [list of transcript metadata]

    if not research_dir.exists():
        logger.warning('Companies directory not found — skipping transcript scan')
        return transcript_manifest

    # Load AI-generated summaries if available
    summaries_file = research_dir / 'transcript_summaries.json'
    ai_summaries = {}
    if summaries_file.exists():
        try:
            with open(summaries_file, 'r') as f:
                ai_summaries = json.load(f)
            logger.info(f'Loaded {len(ai_summaries)} AI transcript summaries')
        except Exception as e:
            logger.warning(f'Could not load transcript summaries: {e}')

    # Find all transcript files
    txt_files = sorted(research_dir.glob('**/Transcripts/*_Earnings_Call.txt'))
    logger.info(f'Found {len(txt_files)} transcript files')

    for fpath in txt_files:
        fname = fpath.stem  # e.g. MG_Q3_2025_Earnings_Call
        parts = fname.split('_')
        if len(parts) < 4:
            continue

        file_ticker = parts[0]
        quarter = parts[1]  # Q1, Q2, Q3, Q4
        year = parts[2]

        # Map to dashboard ticker
        dashboard_ticker = TRANSCRIPT_TICKER_MAP.get(file_ticker, file_ticker)

        # Read transcript
        try:
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
        except Exception as e:
            logger.warning(f'Could not read {fpath}: {e}')
            continue

        # Extract metadata from first lines
        lines = text.split('\n')
        title = lines[0].strip() if lines else fname
        date_str = ''
        company_name = ''
        for line in lines[:6]:
            if line.startswith('Date:'):
                date_str = line.replace('Date:', '').strip()
            if line.startswith('Company:'):
                company_name = line.replace('Company:', '').strip()

        if not company_name:
            company_name = file_ticker

        period = f'{quarter} {year}'
        sort_key = f'{year}-{quarter}'  # for sorting

        # Extract summary
        summary = _extract_summary(text, max_words=250)

        # Generate individual HTML page
        html_filename = f'{file_ticker}_{quarter}_{year}.html'
        html_path = transcripts_out_dir / html_filename
        _generate_transcript_html(text, html_path, company_name, period, date_str)

        # Look up AI summary by key (e.g. MG_Q3_2025)
        summary_key = f'{file_ticker}_{quarter}_{year}'
        ai_summary = ''
        if summary_key in ai_summaries:
            ai_summary = ai_summaries[summary_key].get('summary', '')

        # Build metadata entry
        entry = {
            'period': period,
            'quarter': quarter,
            'year': year,
            'sort_key': sort_key,
            'date': date_str,
            'company': company_name,
            'title': title,
            'summary': ai_summary if ai_summary else summary,
            'html_file': f'transcripts/{html_filename}',
        }

        if dashboard_ticker not in transcript_manifest:
            transcript_manifest[dashboard_ticker] = []
        transcript_manifest[dashboard_ticker].append(entry)

    # Sort each ticker's transcripts newest first
    for ticker in transcript_manifest:
        transcript_manifest[ticker].sort(key=lambda x: x['sort_key'], reverse=True)

    logger.info(f'Generated transcript pages for {len(transcript_manifest)} tickers')
    return transcript_manifest


# =============================================================================
# MAIN
# =============================================================================

def main():
    logger = setup_logging()
    logger.info('=' * 70)
    logger.info('TIC/NDT INSPECTION INTEL DASHBOARD REFRESH')
    logger.info(f'Started: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    logger.info('=' * 70)

    dry_run = '--dry-run' in sys.argv

    try:
        # 1. Fetch yfinance data
        market_data = fetch_all_market_data(logger)
        if not market_data:
            logger.error('No data fetched. Exiting.')
            return

        # 2. Compute peer rankings
        market_data = compute_peer_rankings(market_data, logger)

        # 3. Update price history
        price_history = update_price_history(market_data, logger)

        # 3b. Compute the two basket indices: Inspection-11 (TIC/NDT pure-play) and Flow & MOS-11 (adjacent industrial services)
        inspection11_data = compute_basket_index(price_history, market_data, 'Inspection-11', INSPECTION_11_TICKERS, logger)
        flowmos11_data = compute_basket_index(price_history, market_data, 'Flow & MOS-11', FLOW_MOS_TICKERS, logger)

        # 4. Detect big movers
        big_movers = detect_big_movers(market_data, price_history, logger)

        # 5. Fetch industry indicators from FRED
        industry_indicators = fetch_industry_indicators(logger)

        # 5b. Run Finnhub fetcher for earnings/insider/recommendation data
        try:
            import subprocess
            finnhub_script = SCRIPT_DIR / 'finnhub_fetcher.py'
            if finnhub_script.exists():
                logger.info('Running Finnhub data fetcher...')
                result = subprocess.run(
                    [sys.executable, str(finnhub_script)],
                    capture_output=True, text=True, timeout=120
                )
                if result.returncode == 0:
                    logger.info('Finnhub fetcher completed successfully')
                else:
                    logger.warning(f'Finnhub fetcher returned code {result.returncode}')
                    if result.stderr:
                        logger.warning(f'  stderr: {result.stderr[:200]}')
        except Exception as e:
            logger.warning(f'Finnhub fetcher failed (non-fatal): {e}')

        # 6. Load news data
        news_items = load_news_csv(n=50, logger=logger)

        # 6b. Fetch historical financials for growth trajectories
        historical_financials = fetch_historical_financials(market_data, logger)

        if dry_run:
            logger.info('\n=== DRY RUN — No files written ===')
            logger.info(f'Tickers fetched: {len(market_data)}')
            for t, d in sorted(market_data.items(), key=lambda x: x[1]['market_cap_b'], reverse=True):
                logger.info(f'  {t}: ${d["price_usd"]:.2f} MCap=${d["market_cap_b"]:.1f}B P/E={d.get("pe_trailing", "N/A")} 1Y={d["change_1y"]:+.1f}% Score={d.get("composite_score", "N/A")}')
            logger.info(f'Big movers: {len(big_movers)}')
            logger.info(f'FRED indicators: {len(industry_indicators.get("oil_gas", {}))} O&G, {len(industry_indicators.get("manufacturing", {}))} Mfg')
            logger.info(f'News items: {len(news_items)}')
            return

        # 7. Save JSON files
        DASHBOARD_DIR.mkdir(parents=True, exist_ok=True)
        with open(MARKET_DATA_FILE, 'w') as f:
            json.dump(market_data, f, indent=2, default=str)
        logger.info(f'Saved: {MARKET_DATA_FILE}')

        with open(PRICE_HISTORY_FILE, 'w') as f:
            json.dump(price_history, f, indent=2, default=str)
        logger.info(f'Saved: {PRICE_HISTORY_FILE}')

        with open(INDUSTRY_INDICATORS_FILE, 'w') as f:
            json.dump(industry_indicators, f, indent=2, default=str)
        logger.info(f'Saved: {INDUSTRY_INDICATORS_FILE}')

        # 8. Generate Excel
        generate_excel(market_data, price_history, big_movers, industry_indicators, logger)

        # 9. Inject data into HTML dashboards
        logger.info('Injecting data into HTML dashboards...')

        # Add generated_at timestamp to market_data for all dashboards
        market_data['generated_at'] = datetime.now().isoformat()

        # Equities Dashboard
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Equities_Dashboard.html',
            'INJECTED_MARKET_DATA', market_data, logger)
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Equities_Dashboard.html',
            'INJECTED_HISTORY', historical_financials, logger)

        # Basket indices (Inspection-11 + Flow & MOS-11) — injected only into the home page,
        # which is the sole consumer. The Equities Dashboard's old NDT-10 tab was retired.

        # Peer Analysis Dashboard
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Peer_Analysis_Dashboard.html',
            'INJECTED_MARKET_DATA', market_data, logger)
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Peer_Analysis_Dashboard.html',
            'INJECTED_HISTORY', historical_financials, logger)

        # Load M&A deals data
        ma_deals_path = PROJECT_DIR / 'Industry_Data' / 'ma_deals.json'
        ma_deals = []
        if ma_deals_path.exists():
            with open(ma_deals_path, 'r') as f:
                ma_deals = json.load(f)
            logger.info(f'  Loaded {len(ma_deals)} M&A deals')

        # Industry Dashboard — bundle all live data
        live_data_bundle = {
            'market_data': market_data,
            'big_movers': big_movers,
            'industry_indicators': industry_indicators,
            'ma_deals': ma_deals,
            'generated_at': datetime.now().isoformat(),
        }
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Industry_Dashboard.html',
            'INJECTED_LIVE_DATA', live_data_bundle, logger)

        # Load Finnhub data if available
        finnhub_path = PROJECT_DIR / 'Industry_Data' / 'Finnhub' / 'finnhub_data.json'
        finnhub_data = {}
        if finnhub_path.exists():
            try:
                with open(finnhub_path, 'r', encoding='utf-8') as f:
                    finnhub_data = json.load(f)
                logger.info(f'Loaded Finnhub data: {len(finnhub_data.get("earnings_calendar", []))} earnings, '
                            f'{len(finnhub_data.get("insider_transactions", []))} insider txns, '
                            f'{len(finnhub_data.get("recommendation_trends", {}))} rec trends')
            except Exception as e:
                logger.warning(f'Failed to load Finnhub data: {e}')

        # Command Center Homepage
        cc_ticker_data = {k: v for k, v in market_data.items() if isinstance(v, dict)}
        top_mover = max(cc_ticker_data.values(), key=lambda x: abs(x.get('daily_change_pct') or 0))
        command_center = {
            'market_data': {t: {'price': d.get('price_usd'), 'change_pct': d.get('daily_change_pct') or 0, 'market_cap_b': d.get('market_cap_b'), 'name': d.get('name'), 'display_ticker': d.get('display_ticker', t), 'category': d.get('category')} for t, d in cc_ticker_data.items()},
            'total_mcap': round(sum(d.get('market_cap_b', 0) for d in cc_ticker_data.values()), 1),
            'top_mover': {'ticker': top_mover.get('display_ticker'), 'name': top_mover.get('name'), 'change_pct': top_mover.get('daily_change_pct') or 0},
            'wti': industry_indicators.get('oil_gas', {}).get('dcoilwtico', {}).get('current'),
            'inspection11': inspection11_data.get('current') if inspection11_data else None,
            'inspection11_daily_pct': inspection11_data.get('daily_change_pct') if inspection11_data else None,
            'inspection11_ytd_pct': inspection11_data.get('ytd_pct') if inspection11_data else None,
            'inspection11_index': inspection11_data.get('index', []) if inspection11_data else [],
            'flowmos11': flowmos11_data.get('current') if flowmos11_data else None,
            'flowmos11_daily_pct': flowmos11_data.get('daily_change_pct') if flowmos11_data else None,
            'flowmos11_ytd_pct': flowmos11_data.get('ytd_pct') if flowmos11_data else None,
            'flowmos11_index': flowmos11_data.get('index', []) if flowmos11_data else [],
            'generated_at': datetime.now().isoformat(),
        }
        if finnhub_data.get('earnings_calendar'):
            today_str = datetime.now().strftime('%Y-%m-%d')
            upcoming = [e for e in finnhub_data['earnings_calendar'] if e['date'] >= today_str]
            upcoming.sort(key=lambda x: x['date'])
            if upcoming:
                command_center['next_earnings'] = {'company': upcoming[0].get('company', upcoming[0].get('symbol')), 'date': upcoming[0]['date']}
        # Add transcript analysis data to command center (load from bundle if available)
        try:
            ta_bundle_path = PROJECT_DIR / 'Industry_Data' / 'Transcript_Analysis' / 'analysis_bundle.json'
            if ta_bundle_path.exists():
                with open(ta_bundle_path) as f:
                    command_center['transcript_analysis'] = json.load(f)
                logger.info(f'  Added transcript analysis to Command Center: {len(command_center["transcript_analysis"].get("companies", {}))} companies')
        except Exception as e:
            logger.warning(f'Could not load transcript analysis for Command Center: {e}')
        inject_json_into_html(
            DASHBOARD_DIR / 'index.html',
            'INJECTED_COMMAND_CENTER', command_center, logger)

        # News Dashboard — load all news (larger set than Industry Dashboard)
        all_news_items = load_news_csv(n=500, logger=logger)

        # __analyst_actions_v1 begin — auto-inserted; pulls yfinance upgrades/downgrades
        _analyst_actions = []
        try:
            import sys as _sys
            from pathlib import Path as _Path
            _shared = _Path(__file__).resolve().parent.parent.parent / "_shared"
            if str(_shared) not in _sys.path:
                _sys.path.insert(0, str(_shared))
            from analyst_actions_helper import fetch_analyst_actions
            _aa_info = {t: {'company': d.get('name', t), 'subsector': d.get('category', '')}
                        for t, d in market_data.items() if isinstance(d, dict)}
            _analyst_actions = fetch_analyst_actions(_aa_info, lookback_days=30)
            logger.info(f'  Analyst actions: {len(_analyst_actions)} rows (Inspection_Intel)')
        except Exception as _e:
            logger.warning(f'  Analyst-actions fetch failed for Inspection_Intel: {_e}')
        # __analyst_actions_v1 end

        news_bundle = {
            'news_items': all_news_items,
            'analyst_actions': _analyst_actions,
            'finnhub': finnhub_data,
            'generated_at': datetime.now().isoformat(),
        }
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_News_Dashboard.html',
            'INJECTED_NEWS_DATA', news_bundle, logger)

        # Equities Dashboard — inject Finnhub data
        if finnhub_data:
            inject_json_into_html(
                DASHBOARD_DIR / 'TIC_NDT_Equities_Dashboard.html',
                'INJECTED_FINNHUB', finnhub_data, logger)

        # Auto-summarize any new transcripts before scanning
        try:
            summarizer_script = SCRIPT_DIR / 'summarize_transcripts.py'
            if summarizer_script.exists():
                logger.info('Running transcript auto-summarizer for new transcripts...')
                env = os.environ.copy()
                for k in list(env.keys()):
                    if k.startswith('CLAUDE') or k == 'CLAUDECODE':
                        env.pop(k)
                result = subprocess.run(
                    [sys.executable, str(summarizer_script)],
                    capture_output=True, text=True, timeout=600, env=env
                )
                if result.returncode == 0:
                    logger.info('Transcript summarizer completed successfully')
                    if result.stdout:
                        for line in result.stdout.strip().split('\n')[-5:]:
                            logger.info(f'  {line}')
                else:
                    logger.warning(f'Transcript summarizer returned code {result.returncode}')
                    if result.stderr:
                        logger.warning(f'  stderr: {result.stderr[:300]}')
        except Exception as e:
            logger.warning(f'Transcript summarizer failed (non-fatal): {e}')

        # Run transcript analyzer (sentiment, topics, guidance — fast NLP)
        transcript_analysis = {}
        try:
            analyzer_script = SCRIPT_DIR / 'analyze_transcripts.py'
            if analyzer_script.exists():
                logger.info('Running transcript analyzer (sentiment/topics/guidance)...')
                result = subprocess.run(
                    [sys.executable, str(analyzer_script), '--json-only'],
                    capture_output=True, text=True, timeout=120
                )
                if result.returncode == 0:
                    logger.info('Transcript analyzer completed successfully')
                    if result.stdout:
                        for line in result.stdout.strip().split('\n')[-5:]:
                            logger.info(f'  {line}')
                    # Load the generated JSON bundle
                    bundle_path = PROJECT_DIR / 'Industry_Data' / 'Transcript_Analysis' / 'analysis_bundle.json'
                    if bundle_path.exists():
                        with open(bundle_path) as f:
                            transcript_analysis = json.load(f)
                        logger.info(f'  Loaded analysis bundle: {len(transcript_analysis.get("companies", {}))} companies')
                else:
                    logger.warning(f'Transcript analyzer returned code {result.returncode}')
                    if result.stderr:
                        logger.warning(f'  stderr: {result.stderr[:300]}')
        except Exception as e:
            logger.warning(f'Transcript analyzer failed (non-fatal): {e}')

        # Run transcript intelligence engine (NDT-filtered scorecard, credibility, etc.)
        try:
            intel_script = SCRIPT_DIR / 'transcript_intelligence.py'
            if intel_script.exists():
                logger.info('Running transcript intelligence engine...')
                result = subprocess.run(
                    [sys.executable, str(intel_script)],
                    capture_output=True, text=True, timeout=120
                )
                if result.returncode == 0:
                    logger.info('Transcript intelligence completed successfully')
                else:
                    logger.warning(f'Transcript intelligence returned code {result.returncode}')
                # Load and inject into Equities Dashboard, index.html, and Company Summary
                intel_path = PROJECT_DIR / 'Companies' / 'transcript_intel.json'
                if intel_path.exists():
                    with open(intel_path) as f:
                        intel_data = json.load(f)
                    inject_json_into_html(
                        DASHBOARD_DIR / 'TIC_NDT_Equities_Dashboard.html',
                        'INJECTED_TRANSCRIPT_INTEL', intel_data, logger)
                    inject_json_into_html(
                        DASHBOARD_DIR / 'index.html',
                        'INJECTED_TRANSCRIPT_INTEL', intel_data, logger)
                    inject_json_into_html(
                        DASHBOARD_DIR / 'TIC_NDT_Company_Summary.html',
                        'INJECTED_TRANSCRIPT_INTEL', intel_data, logger)
                    logger.info(f'  Injected transcript intel: scorecard={len(intel_data.get("scorecard", {}))}, credibility={len(intel_data.get("credibility", {}))}')
        except Exception as e:
            logger.warning(f'Transcript intelligence failed (non-fatal): {e}')

        # Company Summary Dashboard — scan transcripts + inject
        transcript_manifest = scan_and_generate_transcripts(logger)

        summary_bundle = {
            'market_data': market_data,
            'historical_financials': historical_financials,
            'finnhub': finnhub_data,
            'transcripts': transcript_manifest,
            'transcript_analysis': transcript_analysis,
            'generated_at': datetime.now().isoformat(),
        }
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Company_Summary.html',
            'INJECTED_SUMMARY_DATA', summary_bundle, logger)

        # 9.5 Scan SEC filings & press releases
        logger.info('Scanning SEC filings & press releases...')
        import re as _re
        sec_filings_manifest = {}
        companies_dir = PROJECT_DIR / 'Companies'
        if companies_dir.exists():
            for cat_dir in sorted(companies_dir.iterdir()):
                if not cat_dir.is_dir() or cat_dir.name.startswith('.'):
                    continue
                for company_dir in sorted(cat_dir.iterdir()):
                    if not company_dir.is_dir():
                        continue
                    # Extract ticker from folder name (e.g., MistrasGroup_MG → MG)
                    parts = company_dir.name.rsplit('_', 1)
                    ticker = parts[-1] if len(parts) > 1 else company_dir.name
                    filings = []
                    for ft in ['10-K', '10-Q', '8-K', 'Transcripts', 'Presentations', 'Press_Releases', 'Supplementals', 'Other']:
                        ft_dir = company_dir / ft
                        if not ft_dir.exists():
                            continue
                        for fpath in sorted(ft_dir.iterdir(), reverse=True):
                            if fpath.is_file() and fpath.suffix in ('.html', '.htm', '.txt', '.pdf'):
                                filings.append({
                                    'type': ft.replace('_', ' '),
                                    'filename': fpath.name,
                                    'path': str(fpath.relative_to(PROJECT_DIR)),
                                    'size_kb': round(fpath.stat().st_size / 1024, 1),
                                    'modified': datetime.fromtimestamp(fpath.stat().st_mtime).strftime('%Y-%m-%d'),
                                })
                    filings.sort(key=lambda x: x.get('modified', ''), reverse=True)
                    if filings:
                        sec_filings_manifest[ticker] = {
                            'company_dir': company_dir.name,
                            'filings': filings[:20],
                            'total_count': len(filings),
                        }
        logger.info(f'  SEC filings scanned: {len(sec_filings_manifest)} companies, '
                    f'{sum(v["total_count"] for v in sec_filings_manifest.values())} total filings')
        inject_json_into_html(
            DASHBOARD_DIR / 'TIC_NDT_Company_Summary.html',
            'SEC_FILINGS', sec_filings_manifest, logger)

        # 10. Summary
        ticker_data = {k: v for k, v in market_data.items() if isinstance(v, dict)}
        total_mcap = sum(d['market_cap_b'] for d in ticker_data.values())
        pe_vals = [d['pe_trailing'] for d in ticker_data.values()
                   if isinstance(d.get('pe_trailing'), (int, float))
                   and not isinstance(d.get('pe_trailing'), bool)]
        best = max(ticker_data.values(), key=lambda x: x.get('change_1y', 0))
        worst = min(ticker_data.values(), key=lambda x: x.get('change_1y', 0))

        logger.info('=' * 70)
        logger.info('SUMMARY')
        logger.info(f'Tickers Updated: {len(ticker_data)}/{len(TICKER_UNIVERSE)}')
        logger.info(f'Total Universe MCap: ${total_mcap:.1f}B')
        if pe_vals:
            logger.info(f'Median P/E (TTM): {statistics.median(pe_vals):.1f}x')
        logger.info(f'Best 1Y: {best["name"]} ({best["change_1y"]:+.1f}%)')
        logger.info(f'Worst 1Y: {worst["name"]} ({worst["change_1y"]:+.1f}%)')
        logger.info(f'Big Movers: {len(big_movers)}')
        logger.info(f'Completed: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        logger.info('=' * 70)

        # ═══════════════════════════════════════════════════
        # EARNINGS DASHBOARD
        # ═══════════════════════════════════════════════════
        earnings_template = DASHBOARD_DIR / 'earnings_template.html'
        earnings_out = DASHBOARD_DIR / 'TIC_NDT_Earnings_Dashboard.html'
        if earnings_template.exists():
            import shutil
            shutil.copy2(earnings_template, earnings_out)

            earnings_calendar = sorted([
                {'ticker': t, 'company': d.get('company', d.get('name', t)),
                 'report_date': d['next_earnings_date'],
                 'est_eps': d.get('eps_forward'),
                 'category': d.get('category', '')}
                for t, d in ticker_data.items() if d.get('next_earnings_date')
            ], key=lambda x: x['report_date'])

            # Load analysis bundle if available
            _analysis_bundle = {}
            _ab_path = PROJECT_DIR / 'Industry_Data' / 'Transcript_Analysis' / 'analysis_bundle.json'
            if _ab_path.exists():
                try:
                    with open(_ab_path) as _abf:
                        _analysis_bundle = json.load(_abf)
                except Exception:
                    pass

            earnings_bundle = {
                'calendar': earnings_calendar,
                'market_data': market_data,
                'transcript_summaries': transcript_manifest,
                'analysis_bundle': _analysis_bundle,
                'category_colors': {},
                'generated_at': datetime.now().isoformat(),
            }
            inject_json_into_html(earnings_out, 'EARNINGS_DATA', earnings_bundle, logger)
            logger.info(f'Earnings Dashboard saved: {earnings_out.name}')

        # ═══════════════════════════════════════════════════
        # AUTO-COMMIT & PUSH TO GITHUB
        # ═══════════════════════════════════════════════════
        if '--dry-run' not in sys.argv:
            try:
                import subprocess
                git_dir = DASHBOARD_DIR
                logger.info('Auto-commit: staging updated dashboard files...')

                # Stage all modified tracked files in Dashboard repo
                files_to_stage = [
                    '*.html', '*.json',
                ]
                for pattern in files_to_stage:
                    subprocess.run(
                        ['git', 'add', pattern],
                        cwd=str(git_dir), capture_output=True, timeout=30
                    )

                # Check if there are staged changes
                status = subprocess.run(
                    ['git', 'diff', '--cached', '--quiet'],
                    cwd=str(git_dir), capture_output=True, timeout=30
                )

                if status.returncode != 0:
                    # There are staged changes — commit and push
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
                    commit_msg = f'Auto-refresh: market data update {timestamp}'
                    result = subprocess.run(
                        ['git', 'commit', '-m', commit_msg],
                        cwd=str(git_dir), capture_output=True, text=True, timeout=60
                    )
                    if result.returncode == 0:
                        logger.info(f'Auto-commit: {result.stdout.strip().split(chr(10))[0]}')

                        push_result = subprocess.run(
                            ['git', 'push'],
                            cwd=str(git_dir), capture_output=True, text=True, timeout=120
                        )
                        if push_result.returncode == 0:
                            logger.info('Auto-push: successfully pushed to GitHub → Cloudflare deploy triggered')
                        else:
                            logger.warning(f'Auto-push failed: {push_result.stderr.strip()}')
                    else:
                        logger.warning(f'Auto-commit failed: {result.stderr.strip()}')
                else:
                    logger.info('Auto-commit: no changes to commit (data unchanged)')

            except Exception as e:
                logger.warning(f'Auto-commit/push failed (non-fatal): {e}')

    except Exception as e:
        logger.error(f'FATAL ERROR: {e}')
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
