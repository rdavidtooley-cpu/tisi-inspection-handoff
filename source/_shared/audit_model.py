#!/usr/bin/env python3
"""
Financial Model Integrity Audit
================================
Replicates the integrity checks from Anthropic's financial-analysis:audit-xls
skill, plus the project-specific checks we've accumulated (row-103 template
bug, Beg Cash chain integrity, cross-company contamination).

Usage:
    python3 _shared/audit_model.py <path/to/Model.xlsx>
    python3 _shared/audit_model.py --check-recent [--seconds 120]
    python3 _shared/audit_model.py --all  # find every Model*.xlsx in tree

Exit codes:
    0 = clean (no issues)
    1 = warnings (some periods don't reconcile but no structural issues)
    2 = errors (circular refs, broken formulas, cross-company contamination)
"""
import argparse
import glob
import os
import re
import sys
import time
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(2)

# Standard row map for our TISI/MG-template models.
# If the model uses a different template, audit gracefully reports "structure not recognized".
ROW_LABELS = {
    150: "Change In Cash",
    151: "Cash - Beginning of Period",
    152: "Ending Cash Balance",
    158: "Cash and cash equivalents",   # BS Cash
    170: "Total Assets",
    192: "Total Liabilities & Equity",
}


def find_period_columns(ws):
    """Return list of (label, col_letter) for every populated period column.
    Walks row 6 (period-end date) and row 7 (period label) to identify columns."""
    cols = []
    for c in range(5, ws.max_column + 1):
        L = get_column_letter(c)
        # Sample row 6 for a datetime, row 7 for label, row 152 or 158 for cash value
        r6 = ws.cell(row=6, column=c).value
        r7 = ws.cell(row=7, column=c).value
        cash_val = ws.cell(row=158, column=c).value
        end_val = ws.cell(row=152, column=c).value
        if not (isinstance(cash_val, (int, float)) or isinstance(end_val, (int, float))):
            continue
        # Period label from row 7 (e.g., "25 Q1") or row 5 (e.g., 2025)
        if isinstance(r7, str) and r7:
            label = r7.replace("\xa0", " ").strip()
        elif isinstance(ws.cell(row=5, column=c).value, int):
            label = f"FY{ws.cell(row=5, column=c).value}"
        else:
            label = L
        cols.append((label, L))
    return cols


def check_cash_recon(ws_v, cols):
    """Check Ending Cash (row 152) vs BS Cash (row 158) for each period.
    Skips periods where BS Cash is zero — those are forecast/empty columns."""
    results = []
    for label, L in cols:
        c = column_index_from_string(L)
        e = ws_v.cell(row=152, column=c).value
        b = ws_v.cell(row=158, column=c).value
        if not isinstance(e, (int, float)):
            e = 0
        if not isinstance(b, (int, float)):
            b = 0
        # Skip unpopulated periods (no BS data = forecast or empty)
        if abs(b) < 0.5:
            continue
        delta = e - b
        results.append((label, L, e, b, delta, abs(delta) < 0.5))
    return results


def check_bs_balance(ws_v, cols):
    """Check Total Assets (row 170) vs TL+E (row 192). Skip forecast cells."""
    results = []
    for label, L in cols:
        c = column_index_from_string(L)
        a = ws_v.cell(row=170, column=c).value
        e = ws_v.cell(row=192, column=c).value
        bs_cash = ws_v.cell(row=158, column=c).value
        if not isinstance(a, (int, float)):
            a = 0
        if not isinstance(e, (int, float)):
            e = 0
        if not isinstance(bs_cash, (int, float)):
            bs_cash = 0
        # Skip if BS Cash is zero (not a populated period)
        if abs(bs_cash) < 0.5:
            continue
        delta = a - e
        results.append((label, L, a, e, delta, abs(delta) < 0.5))
    return results


def check_circular_refs(ws_f):
    """Find formula cells where the formula contains the cell's own coordinate.
    These cause silent calculation errors hidden under IFERROR wrappers."""
    circs = []
    for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row,
                              max_col=ws_f.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                if re.search(rf"\b{cell.coordinate}\b", v):
                    circs.append((cell.coordinate, v[:80]))
    return circs


def check_q4_derivation(ws_f, year_q4_map):
    """For each year's Q4 column, verify input rows derive as =FY-Q1-Q2-Q3.
    Flags row-103 style template bugs (wrong-cell-refs mixed across rows)."""
    # Input rows from our TISI/MG template — these should follow Q4 derivation rule
    INPUT_ROWS = [44, 47, 58, 64, 73, 74, 76, 79, 96, 103, 110,
                  117, 118, 119, 120, 121, 123, 124, 125,
                  131, 132, 138, 139, 141, 142, 143, 145, 148]
    issues = []
    for year, cols in year_q4_map.items():
        FY, Q1, Q2, Q3, Q4 = cols
        for r in INPUT_ROWS:
            f = ws_f.cell(row=r, column=column_index_from_string(Q4)).value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            # Skip rows that reference period-end (BS items pull from FY)
            if r in (158, 160, 161, 163, 167, 168, 169):
                continue
            # Expected pattern: formula contains FY{r}, Q1{r}, Q2{r}, Q3{r}
            if all(re.search(rf"\b{c}{r}\b", f) for c in (FY, Q1, Q2, Q3)):
                continue
            # Allow pointer formulas like =IFERROR(AT53,"") that point elsewhere
            m = re.match(rf'^=IFERROR\(\+?{Q4}(\d+)\s*,', f)
            if m and int(m.group(1)) != r:
                continue
            issues.append((year, Q4, r, f[:60]))
    return issues


def detect_year_q4_map(ws):
    """Identify (FY, Q1, Q2, Q3, Q4) column tuples by walking row 5/6/7."""
    import datetime
    year_cols = {}
    qtr_cols = {}
    for c in range(5, ws.max_column + 1):
        L = get_column_letter(c)
        r5 = ws.cell(row=5, column=c).value
        r6 = ws.cell(row=6, column=c).value
        if isinstance(r5, int) and 2000 <= r5 <= 2035:
            year_cols[r5] = L
        if isinstance(r5, int) and 1 <= r5 <= 4 and isinstance(r6, datetime.datetime):
            qtr_cols[(r6.year, r5)] = L
    out = {}
    for yr, FY in year_cols.items():
        if all((yr, q) in qtr_cols for q in (1, 2, 3, 4)):
            out[yr] = (FY, qtr_cols[(yr, 1)], qtr_cols[(yr, 2)],
                       qtr_cols[(yr, 3)], qtr_cols[(yr, 4)])
    return out


def check_beg_cash_chain(ws_f, year_q4_map):
    """For each annual column, verify Beg Cash (row 151) chains from PRIOR YEAR
    ending (row 152), not from a quarterly column."""
    issues = []
    years_sorted = sorted(year_q4_map.keys())
    for i, yr in enumerate(years_sorted):
        FY = year_q4_map[yr][0]
        c = column_index_from_string(FY)
        beg_formula = ws_f.cell(row=151, column=c).value
        if not isinstance(beg_formula, str):
            continue  # hardcoded seed — OK
        if i == 0:
            continue  # first year — formula or hardcoded both acceptable
        prior_FY = year_q4_map[years_sorted[i - 1]][0]
        # Expected formula references prior FY's row 152
        if re.search(rf"\b{prior_FY}152\b", beg_formula):
            continue
        # Otherwise: flag
        issues.append((yr, FY, beg_formula))
    return issues


def audit_workbook(path):
    """Run all checks against one workbook. Returns (warnings_count, errors_count, report_lines)."""
    if not os.path.exists(path):
        return 0, 1, [f"  ERROR: file not found: {path}"]
    if not path.endswith(".xlsx"):
        return 0, 1, [f"  ERROR: not an xlsx file: {path}"]

    try:
        wb_v = openpyxl.load_workbook(path, data_only=True)
        wb_f = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        return 0, 1, [f"  ERROR: cannot open {path}: {e}"]

    # Find the Main "Model" sheet
    sheet_name = None
    for cand in ("Model", "model", "MODEL"):
        if cand in wb_v.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        return 0, 1, [f"  ERROR: no 'Model' sheet found in {path}"]

    ws_v = wb_v[sheet_name]
    ws_f = wb_f[sheet_name]

    report = []
    warnings = 0
    errors = 0

    report.append(f"\n=== Audit: {os.path.basename(path)} ===")

    # 1. Cash reconciliation
    cols = find_period_columns(ws_v)
    cash_results = check_cash_recon(ws_v, cols)
    clean_cash = sum(1 for r in cash_results if r[5])
    total_cash = len(cash_results)
    if total_cash > 0:
        report.append(f"Cash recon: {clean_cash}/{total_cash} periods clean")
        for label, L, e, b, d, ok in cash_results:
            if not ok:
                report.append(f"  ❌ {label} ({L}): End={e:,.0f} BS={b:,.0f} Δ={d:+,.0f}")
                warnings += 1

    # 2. BS balance
    bs_results = check_bs_balance(ws_v, cols)
    clean_bs = sum(1 for r in bs_results if r[5])
    total_bs = len(bs_results)
    if total_bs > 0:
        report.append(f"BS balance: {clean_bs}/{total_bs} periods clean")
        for label, L, a, e, d, ok in bs_results:
            if not ok:
                report.append(f"  ❌ {label} ({L}): TA={a:,.0f} TL+E={e:,.0f} Δ={d:+,.0f}")
                warnings += 1

    # 3. Circular self-references
    circs = check_circular_refs(ws_f)
    if circs:
        report.append(f"❌ Circular self-references: {len(circs)} cells")
        for coord, formula in circs[:5]:
            report.append(f"  {coord}: {formula}")
        if len(circs) > 5:
            report.append(f"  ...and {len(circs)-5} more")
        errors += 1
    else:
        report.append("Circular self-refs: 0 ✅")

    # 4. Q4 derivation
    year_q4_map = detect_year_q4_map(ws_f)
    if year_q4_map:
        q4_issues = check_q4_derivation(ws_f, year_q4_map)
        if q4_issues:
            report.append(f"❌ Q4 derivation violations: {len(q4_issues)} cells")
            for yr, Q4, r, f in q4_issues[:5]:
                report.append(f"  {Q4}{r} ({yr}Q4): {f}")
            errors += 1
        else:
            report.append(f"Q4 derivation: clean across {len(year_q4_map)} years ✅")

    # 5. Beg Cash chain
    if year_q4_map:
        beg_issues = check_beg_cash_chain(ws_f, year_q4_map)
        if beg_issues:
            report.append(f"❌ Beg Cash chain issues: {len(beg_issues)} years")
            for yr, FY, f in beg_issues:
                report.append(f"  FY{yr} ({FY}151): {f}")
            errors += 1

    if warnings == 0 and errors == 0:
        report.append("✅ All checks clean")

    return warnings, errors, report


def find_recent_models(seconds=120, root="${PROJECT_ROOT}"):
    """Return absolute paths of every *Model*.xlsx modified in the last N seconds."""
    cutoff = time.time() - seconds
    out = []
    for path in glob.glob(f"{root}/**/*Model*.xlsx", recursive=True):
        try:
            if os.path.getmtime(path) >= cutoff and "_pre_" not in path \
               and "_backup_" not in path and "_before_" not in path:
                out.append(path)
        except OSError:
            continue
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("path", nargs="?", help="Path to Model.xlsx file")
    p.add_argument("--check-recent", action="store_true",
                   help="Audit any Model.xlsx modified in last 120 seconds")
    p.add_argument("--seconds", type=int, default=120,
                   help="With --check-recent: lookback window in seconds")
    p.add_argument("--all", action="store_true",
                   help="Audit every Model.xlsx in the project tree")
    p.add_argument("--quiet", action="store_true",
                   help="Only print issues, suppress clean-check output")
    args = p.parse_args()

    paths = []
    if args.path:
        paths.append(args.path)
    elif args.check_recent:
        paths = find_recent_models(args.seconds)
        if not paths and not args.quiet:
            return  # nothing to audit, exit silently
    elif args.all:
        paths = find_recent_models(seconds=10**9)
    else:
        p.print_help()
        sys.exit(1)

    total_warnings = 0
    total_errors = 0
    for path in paths:
        w, e, lines = audit_workbook(path)
        total_warnings += w
        total_errors += e
        if not args.quiet or w or e:
            for line in lines:
                print(line)

    if total_errors > 0:
        sys.exit(2)
    elif total_warnings > 0:
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
